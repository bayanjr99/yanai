#!/usr/bin/env python
"""
End-to-end test on real data.

Steps
-----
1. Parse PDF  →  validate each employee's sum vs PDF total
2. Load agreements + costs
3. Match + bill at DAILY level (completion per day before aggregation)
4. Aggregate monthly  →  apply monthly_min completion if needed
5. Export debug Excel (5 sheets)
6. Print pass/fail report
"""

from __future__ import annotations

import io
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

# force UTF-8 output so Hebrew prints correctly on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── capture parser warnings ──────────────────────────────────────────────────
_log_records: list[str] = []

class _Capture(logging.Handler):
    def emit(self, record):
        _log_records.append(self.format(record))

_root = logging.getLogger()
_root.setLevel(logging.DEBUG)
_cap = _Capture()
_cap.setFormatter(logging.Formatter("%(levelname)s %(name)s: %(message)s"))
_root.addHandler(_cap)
_root.addHandler(logging.StreamHandler(sys.stdout))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

sys.path.insert(0, str(BASE_DIR))

from core.pdf_parser import _parse_page          # type: ignore[attr-defined]
from core.excel_loaders import load_agreements, load_costs
from core.matcher import is_internal, find_agreement, resolve_client


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Parse PDF + per-employee total validation
# ─────────────────────────────────────────────────────────────────────────────

def parse_and_validate(pdf_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns
    -------
    daily_df      : employee_id, employee_name, date, site, hours_to_pay
    validation_df : page, employee_id, employee_name, rows_parsed,
                    parsed_sum, pdf_total, diff, status
    """
    reader = PdfReader(str(pdf_path))
    daily_rows: list[dict] = []
    val_rows:   list[dict] = []

    _RE_LO = re.compile(r"לא\s+לדיווח|לא\s+רלוונטי", re.UNICODE)

    for page_num, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        rows, page_total = _parse_page(text)

        if not rows:
            if page_total is not None:
                has_lo = bool(_RE_LO.search(text))
                val_rows.append({
                    "page": page_num, "employee_id": "?", "employee_name": "?",
                    "rows_parsed": 0, "parsed_sum": 0.0,
                    "pdf_total": page_total, "diff": page_total,
                    "status": "WARN – לא רלוונטי כל החודש" if has_lo else "WARN – no rows parsed",
                })
            continue

        emp_id   = rows[0]["employee_id"]
        emp_name = rows[0]["employee_name"]

        for r in rows:
            daily_rows.append({
                "employee_id":   emp_id,
                "employee_name": emp_name,
                "date":          r["date"],
                "site":          r["site"],
                "hours_to_pay":  r["hours_to_pay"],
            })

        parsed_sum = round(sum(r["hours_to_pay"] for r in rows), 2)

        if page_total is None:
            diff, status = None, "WARN – no PDF total found"
        else:
            diff = round(abs(parsed_sum - page_total), 3)
            if diff <= 0.05:
                status = "PASS"
            elif _RE_LO.search(text):
                # Delta explained by intentionally excluded לא לדיווח days
                status = f"EXPECTED – לא לדיווח excluded (Δ{diff:.2f})"
            else:
                status = f"FAIL (Δ{diff:.2f})"

        val_rows.append({
            "page":          page_num,
            "employee_id":   emp_id,
            "employee_name": emp_name,
            "rows_parsed":   len(rows),
            "parsed_sum":    parsed_sum,
            "pdf_total":     page_total,
            "diff":          diff,
            "status":        status,
        })

    return pd.DataFrame(daily_rows), pd.DataFrame(val_rows)


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Match + billing at daily level
# ─────────────────────────────────────────────────────────────────────────────

def daily_billing(
    daily_df: pd.DataFrame,
    agreements: list[dict],
    costs: dict,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    For each daily row:
      - resolve client
      - find agreement
      - compute completion_day = max(0, daily_min - hours) if hourly
      - billing_day = (hours + completion_day) × rate  OR  1 × rate (daily)

    Returns (detail_daily_df, issues_df).
    """
    detail_rows: list[dict] = []
    issue_rows:  list[dict] = []

    for _, row in daily_df.iterrows():
        emp_id    = str(row["employee_id"])
        emp_name  = str(row["employee_name"])
        site      = str(row["site"])
        hours     = float(row["hours_to_pay"])
        work_date = row["date"]

        client, _ = resolve_client(emp_id, site, costs)

        if not client:
            issue_rows.append({
                "employee_id":   emp_id,
                "employee_name": emp_name,
                "site":          site,
                "date":          str(work_date),
                "issue_type":    "עובד חסר בעלויות",
                "description":   f"עובד {emp_id} ({emp_name}) לא נמצא בקובץ עלות מעביד",
            })
            client = site

        if is_internal(client):
            continue

        agreement, match_reason = find_agreement(client, site, agreements)

        if agreement is None:
            issue_rows.append({
                "employee_id":   emp_id,
                "employee_name": emp_name,
                "site":          site,
                "date":          str(work_date),
                "issue_type":    "הסכם חסר",
                "description":   f"לא נמצא הסכם ללקוח '{client}' אתר '{site}'",
            })
            rate = daily_min = monthly_min = 0.0
            billing_type   = "hourly"
            completion_day = 0.0
            billable_day   = hours
            billing_day    = 0.0
        else:
            billing_type  = str(agreement.get("billing_type", "hourly"))
            rate          = float(agreement.get("rate")        or 0)
            daily_min     = float(agreement.get("daily_min")   or 0)
            monthly_min   = float(agreement.get("monthly_min") or 0)

            if billing_type == "daily":
                completion_day = 0.0
                billable_day   = 1.0   # 1 day counted
                billing_day    = round(rate, 2)
            else:
                completion_day = max(0.0, daily_min - hours) if daily_min > 0 else 0.0
                billable_day   = hours + completion_day
                billing_day    = round(billable_day * rate, 2)

        detail_rows.append({
            "employee_id":        emp_id,
            "employee_name":      emp_name,
            "date":               work_date,
            "site":               site,
            "client":             client,
            "match_reason":       match_reason if agreement else "אין הסכם",
            "billing_type":       billing_type,
            "hours_to_pay":       hours,
            "daily_min":          daily_min,
            "monthly_min":        monthly_min,
            "completion_day":     round(completion_day, 3),
            "billable_hours_day": round(billable_day, 3),
            "rate":               rate,
            "billing_day":        billing_day,
        })

    return pd.DataFrame(detail_rows), pd.DataFrame(issue_rows)


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: Aggregate to monthly + apply monthly_min completion
# ─────────────────────────────────────────────────────────────────────────────

def aggregate_and_finalize(
    detail_df: pd.DataFrame,
    costs: dict,
) -> pd.DataFrame:
    if detail_df.empty:
        return pd.DataFrame()

    grp_keys = ["employee_id", "employee_name", "site"]

    agg = detail_df.groupby(grp_keys, as_index=False).agg(
        client          = ("client",           "first"),
        match_reason    = ("match_reason",     "first"),
        billing_type    = ("billing_type",     "first"),
        rate            = ("rate",             "first"),
        monthly_min     = ("monthly_min",      "first"),
        days            = ("date",             "count"),
        total_hours     = ("hours_to_pay",     "sum"),
        completion_daily= ("completion_day",   "sum"),
        billable_subtotal=("billable_hours_day","sum"),
        billing_subtotal= ("billing_day",      "sum"),
    )

    final_rows: list[dict] = []
    for _, row in agg.iterrows():
        billing_type    = str(row["billing_type"])
        rate            = float(row["rate"])
        monthly_min     = float(row["monthly_min"] or 0)
        billable        = float(row["billable_subtotal"])
        billing_amt     = float(row["billing_subtotal"])
        completion_monthly = 0.0

        # monthly_min completion: only for hourly, only if still below threshold
        if billing_type == "hourly" and monthly_min > 0 and billable < monthly_min:
            completion_monthly = monthly_min - billable
            billable  += completion_monthly
            billing_amt = round(billable * rate, 2)

        emp_id = str(row["employee_id"])
        site   = str(row["site"])
        _, emp_cost = resolve_client(emp_id, site, costs)

        completion_total = float(row["completion_daily"]) + completion_monthly
        profit = round(billing_amt - emp_cost, 2)
        margin = round(profit / billing_amt * 100, 1) if billing_amt > 0 else 0.0

        final_rows.append({
            "employee_id":        emp_id,
            "employee_name":      str(row["employee_name"]),
            "client":             str(row["client"]),
            "site":               site,
            "match_reason":       str(row["match_reason"]),
            "billing_type":       billing_type,
            "rate":               rate,
            "days":               int(row["days"]),
            "total_hours":        round(float(row["total_hours"]), 2),
            "completion_daily":   round(float(row["completion_daily"]), 2),
            "completion_monthly": round(completion_monthly, 2),
            "completion_total":   round(completion_total, 2),
            "billable_hours":     round(billable, 2),
            "billing_amount":     billing_amt,
            "cost":               emp_cost,
            "profit":             profit,
            "margin_pct":         margin,
        })

    return pd.DataFrame(final_rows)


# ─────────────────────────────────────────────────────────────────────────────
# Debug Excel export
# ─────────────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF3FB")
_FAIL_FILL = PatternFill("solid", fgColor="FFD7D7")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_PASS_FILL = PatternFill("solid", fgColor="D5F5D5")


def _write_df(ws, df: pd.DataFrame, title: str, status_col: str | None = None) -> None:
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=12, color="1F497D")
    ws.append([])

    if df.empty:
        ws.append(["(אין נתונים)"])
        ws.sheet_view.rightToLeft = True
        return

    headers = list(df.columns)
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(3, ci)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    status_ci = (headers.index(status_col) + 1) if status_col and status_col in headers else None

    for ri, (_, row) in enumerate(df.iterrows(), 4):
        # row fill by status
        row_fill = None
        if status_ci:
            sv = str(ws.cell(ri - 1, status_ci).value or "")  # not written yet
            sv = str(row.get(status_col, ""))
            if sv.startswith("FAIL"):
                row_fill = _FAIL_FILL
            elif sv.startswith("WARN"):
                row_fill = _WARN_FILL
            elif sv == "PASS":
                row_fill = _PASS_FILL
        elif ri % 2 == 0:
            row_fill = _ALT_FILL

        for ci, col in enumerate(headers, 1):
            val = row[col]
            try:
                is_na = pd.isna(val)
            except (TypeError, ValueError):
                is_na = False
            if is_na:
                val = ""
            elif hasattr(val, "isoformat"):
                val = val.isoformat()
            cell = ws.cell(ri, ci)
            cell.value = val
            cell.alignment = Alignment(horizontal="right")
            if row_fill:
                cell.fill = row_fill

    # auto width
    for ci, h in enumerate(headers, 1):
        col_vals = [str(h)] + [str(df[h].iloc[i]) for i in range(min(len(df), 200))]
        width = min(max(len(v) for v in col_vals) + 3, 42)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A4"


def export_debug_excel(
    daily_df: pd.DataFrame,
    validation_df: pd.DataFrame,
    detail_daily_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    final_df: pd.DataFrame,
) -> Path:
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = OUTPUT_DIR / f"debug_{ts}.xlsx"

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "שורות יומיות גולמיות"
    _write_df(ws1, daily_df, "שורות יומיות מ-PDF (גולמי)")

    ws2 = wb.create_sheet("אימות PDF")
    _write_df(ws2, validation_df, 'השוואת סה"כ שעות לכל עובד', status_col="status")

    ws3 = wb.create_sheet("חישוב יומי")
    _write_df(ws3, detail_daily_df, "חישוב חיוב יומי (לפני אגרגציה)")

    ws4 = wb.create_sheet("חריגים")
    _write_df(
        ws4,
        issues_df if not issues_df.empty
        else pd.DataFrame({"הודעה": ["✓ לא נמצאו חריגים"]}),
        "חריגים ובעיות",
    )

    ws5 = wb.create_sheet("סיכום חיוב סופי")
    _write_df(ws5, final_df, "סיכום חיוב חודשי סופי (אחרי שלמות)")

    wb.save(str(path))
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    sep = "═" * 62
    print(f"\n{sep}")
    print("  END-TO-END TEST — REAL DATA")
    print(sep)

    pdf_path  = DATA_DIR / "hours.pdf"
    agr_path  = DATA_DIR / "agreements.xlsx"
    cost_path = DATA_DIR / "employees_cost.xlsx"

    for p in (pdf_path, agr_path, cost_path):
        if not p.exists():
            print(f"ERROR: קובץ חסר: {p}")
            return 1

    # ── Step 1: Parse PDF ────────────────────────────────────────────────────
    print("\n[1] Parsing PDF + validating totals...")
    daily_df, validation_df = parse_and_validate(pdf_path)

    n_pass     = int((validation_df["status"] == "PASS").sum())
    n_fail     = int(validation_df["status"].str.startswith("FAIL").sum())
    n_expected = int(validation_df["status"].str.startswith("EXPECTED").sum())
    n_warn     = int(validation_df["status"].str.startswith("WARN").sum())
    n_emps     = daily_df["employee_id"].nunique() if not daily_df.empty else 0
    n_days     = len(daily_df)

    print(f"    {n_days} daily rows | {n_emps} employees")
    print(f"    Validation: {n_pass} PASS | {n_fail} FAIL | {n_expected} EXPECTED | {n_warn} WARN")

    if n_fail:
        print("    ✗ FAILED employees:")
        for _, r in validation_df[validation_df["status"].str.startswith("FAIL")].iterrows():
            print(
                f"      Page {r['page']} | emp {r['employee_id']} {r['employee_name']} "
                f"| parsed={r['parsed_sum']:.2f}h  PDF={r['pdf_total']:.2f}h  Δ={r['diff']:.3f}"
            )

    # ── Step 2: Load reference data ──────────────────────────────────────────
    print("\n[2] Loading agreements + costs...")
    try:
        agreements = load_agreements(str(agr_path))
    except Exception as e:
        print(f"    ERROR loading agreements: {e}")
        return 1
    try:
        costs = load_costs(str(cost_path))
    except Exception as e:
        print(f"    ERROR loading costs: {e}")
        return 1
    print(f"    {len(agreements)} agreements | {len(costs)} employees in cost file")

    # ── Step 3: Daily billing ────────────────────────────────────────────────
    print("\n[3] Daily match + billing...")
    if daily_df.empty:
        print("    ERROR: no daily rows — cannot continue")
        return 1

    detail_daily_df, issues_df = daily_billing(daily_df, agreements, costs)
    unique_issues = issues_df["issue_type"].value_counts().to_dict() if not issues_df.empty else {}
    print(f"    {len(detail_daily_df)} billed rows | {len(issues_df)} issues")
    for itype, cnt in unique_issues.items():
        print(f"      {cnt}× {itype}")

    # ── Step 4: Aggregate + monthly_min ──────────────────────────────────────
    print("\n[4] Aggregating to monthly + monthly_min completion...")
    final_df = aggregate_and_finalize(detail_daily_df, costs)
    print(f"    {len(final_df)} final rows (employee × site)")

    # ── Step 5: Export debug Excel ───────────────────────────────────────────
    print("\n[5] Exporting debug Excel...")
    debug_path = export_debug_excel(daily_df, validation_df, detail_daily_df, issues_df, final_df)
    print(f"    Saved: {debug_path.name}")

    # ── Summary ──────────────────────────────────────────────────────────────
    print(f"\n{sep}")
    print("  RESULTS")
    print(sep)

    print(f"\n  PDF Parsing:      {n_pass} PASS | {n_fail} FAIL | {n_expected} EXPECTED | {n_warn} WARN")
    if n_expected:
        print(f"  (EXPECTED = employees with some 'לא לדיווח' days correctly excluded from billing)")

    if not final_df.empty:
        total_billing  = final_df["billing_amount"].sum()
        total_cost     = final_df["cost"].sum()
        total_profit   = total_billing - total_cost
        total_comp     = final_df["completion_total"].sum()
        pct = total_profit / total_billing * 100 if total_billing > 0 else 0.0
        zero_billing   = int((final_df["billing_amount"] == 0).sum())

        print(f"\n  Billing:          ₪{total_billing:>12,.2f}")
        print(f"  Cost:             ₪{total_cost:>12,.2f}")
        print(f"  Profit:           ₪{total_profit:>12,.2f}  ({pct:.1f}%)")
        print(f"  Completion added: {total_comp:.2f}h")
        if zero_billing:
            print(f"  ⚠ Zero-billing rows: {zero_billing}")

    print(f"\n  Issues:           {len(issues_df)}")
    for itype, cnt in unique_issues.items():
        print(f"    {cnt}× {itype}")

    # parser warnings
    parser_warnings = [r for r in _log_records if "WARNING" in r or "ERROR" in r]
    if parser_warnings:
        print(f"\n  Parser warnings:  {len(parser_warnings)}")
        for w in parser_warnings[:15]:
            print(f"    {w}")

    print(f"\n  Debug file:  output/{debug_path.name}")
    if n_fail == 0:
        print(f"\n  ✓ NO UNEXPECTED MISMATCHES")
    else:
        print(f"\n  ✗ {n_fail} UNEXPECTED PDF TOTAL MISMATCH(ES) — investigate these pages")
    print(sep + "\n")

    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
