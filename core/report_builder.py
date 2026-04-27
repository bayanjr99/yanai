"""
Build and save the final Excel reports.

billing report  (final_YYYYMMDD_HHMMSS.xlsx)
  Sheet 1 – "סיכום לפי לקוח"   : one row per client+site with totals
  Sheet 2 – "פירוט לפי עובד"   : one row per employee+site
  Sheet 3 – "ניתוח רווחיות"    : profitability per client

issues report   (issues_YYYYMMDD_HHMMSS.xlsx)
  Sheet 1 – "חריגים"           : all data-quality issues
"""

from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column label maps  (internal key → Hebrew label)
# ---------------------------------------------------------------------------

# Main billing report — matches the client-facing reference format
_BILLING_REPORT_MAP = {
    "client":              "שם לקוח",
    "site":                "אתר",
    "billing_type_label":  "סוג חיוב",
    "rate":                "תעריף ₪",
    "standard":            "תקן",
    "hours_actual":        "כמות שעות",
    "days_actual":         "כמות ימים",
    "billing_actual":      "חיוב על שעות בפועל",
    "completion_hours":    "כמות שעות השלמה",
    "billing_completion":  "חיוב על שעות השלמה",
    "billing_total":       'סה"כ חיוב לפני מע"מ',
}

_SUMMARY_MAP = {
    "client":           "לקוח",
    "site":             "אתר",
    "employees":        "עובדים",
    "days":             "ימי עבודה",
    "total_hours":      "שעות עבודה",
    "billable_hours":   "שעות לחיוב",
    "completion_added": "שלמות",
    "billing_amount":   "סכום לחיוב ₪",
    "cost":             "עלות מעביד ₪",
    "profit":           "רווח ₪",
    "margin_pct":       "% רווח",
}

_DETAIL_MAP = {
    "employee_id":      "מס' עובד",
    "employee_name":    "שם עובד",
    "client":           "לקוח",
    "site":             "אתר",
    "match_reason":     "אופן התאמה",
    "days":             "ימים",
    "total_hours":      "שעות לתשלום",
    "break_hours":      "שעות הפסקה",
    "billable_hours":   "שעות לחיוב",
    "completion_added": "שלמות שנוספה",
    "billing_amount":   "חיוב ₪",
    "cost":             "עלות מעביד ₪",
    "profit":           "רווח ₪",
    "margin_pct":       "% רווח",
}

_PROFIT_MAP = {
    "client":         "לקוח",
    "billing_amount": "סכום חיוב ₪",
    "cost":           "עלות מעביד ₪",
    "profit":         "רווח ₪",
    "margin_pct":     "% רווח",
    "employees":      "מספר עובדים",
}

_ISSUES_MAP = {
    "employee_id":   "מס' עובד",
    "employee_name": "שם עובד",
    "site":          "אתר",
    "issue_type":    "סוג בעיה",
    "description":   "תיאור",
}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F497D")
_ALT_FILL     = PatternFill("solid", fgColor="EEF3FB")
_TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
_WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
_ERROR_FILL   = PatternFill("solid", fgColor="FFE0E0")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FONT   = Font(bold=True, size=10)
_THIN_BORDER  = Border(
    bottom=Side(style="thin", color="AAAAAA"),
)

_MONEY_FMT    = '#,##0.00 ₪'
_PCT_FMT      = '0.0"%"'
_INT_FMT      = '#,##0'


def _write_sheet(
    ws,
    df: pd.DataFrame,
    col_map: dict[str, str],
    title: str,
    add_totals: bool = False,
) -> None:
    """Write a DataFrame to a worksheet with headers, alternating rows, and optional totals."""
    ordered_cols = [c for c in col_map if c in df.columns]
    headers      = [col_map[c] for c in ordered_cols]

    # Title row
    ws.append([title])
    title_cell = ws.cell(1, 1)
    title_cell.font      = Font(bold=True, size=13, color="1F497D")
    title_cell.alignment = Alignment(horizontal="right")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws.append([])  # spacer

    # Header row (row 3)
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(3, col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        for col_idx, col in enumerate(ordered_cols, 1):
            val  = row[col]
            cell = ws.cell(row_idx, col_idx)
            cell.value     = None if pd.isna(val) else val
            cell.alignment = Alignment(horizontal="right")
            if row_idx % 2 == 0:
                cell.fill = _ALT_FILL

            # Number formats
            if col in ("billing_amount", "cost", "profit"):
                cell.number_format = _MONEY_FMT
            elif col == "margin_pct":
                cell.number_format = _PCT_FMT
            elif col in ("days", "employees"):
                cell.number_format = _INT_FMT

    # Totals row
    if add_totals and not df.empty:
        total_row_idx = 4 + len(df)
        ws.cell(total_row_idx, 1).value = "סה\"כ"
        ws.cell(total_row_idx, 1).font  = _TOTAL_FONT
        for col_idx, col in enumerate(ordered_cols, 1):
            cell = ws.cell(total_row_idx, col_idx)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            if col in ("billing_amount", "cost", "profit"):
                total = df[col].sum()
                cell.value         = total
                cell.number_format = _MONEY_FMT
            elif col in ("days", "total_hours", "billable_hours", "completion_added"):
                cell.value         = df[col].sum()
                cell.number_format = _INT_FMT

    # Auto column widths
    for col_idx, col in enumerate(ordered_cols, 1):
        header_len = len(headers[col_idx - 1])
        if not df.empty:
            data_len = df[col].astype(str).str.len().max()
        else:
            data_len = 0
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(header_len, data_len) + 4, 45
        )

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Summary builders
# ---------------------------------------------------------------------------

_COMPLETION_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow — completion cols
_DAILY_FILL      = PatternFill("solid", fgColor="E2EFDA")   # green — days col
_TOTAL_COL_FILL  = PatternFill("solid", fgColor="D6E4F0")   # blue  — total col


def _style_billing_report(ws, df: pd.DataFrame) -> None:
    """Extra styling on the billing report sheet after _write_sheet writes it."""
    if df.empty:
        return
    cols = list(_BILLING_REPORT_MAP.keys())
    comp_idx  = [i + 1 for i, c in enumerate(cols) if c in ("completion_hours", "billing_completion")]
    days_idx  = [i + 1 for i, c in enumerate(cols) if c == "days_actual"]
    total_idx = [i + 1 for i, c in enumerate(cols) if c == "billing_total"]

    n_rows = 4 + len(df)   # header is row 3, data starts row 4
    for ri in range(3, n_rows + 2):   # include totals row
        for ci in comp_idx:
            cell = ws.cell(ri, ci)
            if ri == 3:
                cell.fill = PatternFill("solid", fgColor="BF8F00")
            elif cell.fill.fgColor.rgb == "00000000":   # only if not already coloured
                cell.fill = _COMPLETION_FILL
        for ci in days_idx:
            cell = ws.cell(ri, ci)
            if ri != 3 and cell.fill.fgColor.rgb == "00000000":
                cell.fill = _DAILY_FILL
        for ci in total_idx:
            cell = ws.cell(ri, ci)
            if ri >= 4:
                cell.font = Font(bold=True, size=10)


def _build_billing_report(detail: pd.DataFrame) -> pd.DataFrame:
    """
    Build the client-facing billing report (matches reference Excel format).

    One row per (client, site).
    For hourly billing: fill כמות שעות; כמות ימים stays blank.
    For daily billing:  fill כמות ימים;  כמות שעות shown for reference.
    Billing = actual + completion; all before VAT.
    """
    if detail.empty:
        return pd.DataFrame(columns=list(_BILLING_REPORT_MAP.keys()))

    has_type = "billing_type" in detail.columns
    has_rate = "rate"         in detail.columns
    has_ot   = "ot_hours"     in detail.columns
    has_otr  = "ot_rate"      in detail.columns
    has_otth = "ot_threshold" in detail.columns
    has_mm   = "monthly_min"  in detail.columns

    rows = []
    for (client, site), grp in detail.groupby(["client", "site"], sort=False):
        billing_type = str(grp["billing_type"].iloc[0])    if has_type else "hourly"
        rate         = float(grp["rate"].iloc[0])          if has_rate else 0.0
        ot_rate      = float(grp["ot_rate"].iloc[0])       if has_otr  else 0.0
        ot_threshold = float(grp["ot_threshold"].iloc[0])  if has_otth else 10.0
        mm           = float(grp["monthly_min"].iloc[0])   if has_mm   else 0.0

        total_hours = round(float(grp["total_hours"].sum()),    2)
        ot_hours    = round(float(grp["ot_hours"].sum()),       2) if has_ot else 0.0
        days        = int(grp["days"].sum())
        completion  = round(float(grp["completion_added"].sum()), 2)
        billing_tot = round(float(grp["billing_amount"].sum()),   2)

        if billing_type == "daily_plus_ot":
            billing_actual     = round(days * rate + ot_hours * ot_rate, 2)
            billing_completion = 0.0
            hours_col          = ot_hours if ot_hours > 0 else None   # OT hours
            days_col           = float(days)
            type_label         = "יומי+שע"
            rate_display       = f"{int(rate)}/{int(ot_rate)}"        # e.g. "800/80"
        elif billing_type == "daily":
            billing_actual     = round(days * rate, 2)
            billing_completion = 0.0
            hours_col          = total_hours    # shown for reference
            days_col           = float(days)
            type_label         = "יומי"
            rate_display       = rate
        else:
            billing_actual     = round(total_hours * rate, 2)
            billing_completion = round(completion * rate, 2)
            hours_col          = total_hours
            days_col           = None
            type_label         = "שעתי"
            rate_display       = rate

        rows.append({
            "client":             client,
            "site":               site,
            "billing_type_label": type_label,
            "rate":               rate_display,
            "standard":           mm if mm > 0 else None,
            "hours_actual":       hours_col,
            "days_actual":        days_col,
            "billing_actual":     billing_actual,
            "completion_hours":   completion if completion > 0 else None,
            "billing_completion": billing_completion if billing_completion > 0 else None,
            "billing_total":      billing_tot,
        })

    return pd.DataFrame(rows)


def _build_summary(detail: pd.DataFrame) -> pd.DataFrame:
    if detail.empty:
        return pd.DataFrame(columns=list(_SUMMARY_MAP.keys()))

    grp = (
        detail.groupby(["client", "site"], as_index=False)
        .agg(
            employees      =("employee_id", "nunique"),
            days           =("days", "sum"),
            total_hours    =("total_hours", "sum"),
            billable_hours =("billable_hours", "sum"),
            completion_added=("completion_added", "sum"),
            billing_amount =("billing_amount", "sum"),
            cost           =("cost", "sum"),
        )
    )
    grp["profit"]     = grp["billing_amount"] - grp["cost"]
    grp["margin_pct"] = (
        grp["profit"] / grp["billing_amount"].replace(0, float("nan")) * 100
    ).round(1)
    return grp


def _build_profitability(detail: pd.DataFrame) -> pd.DataFrame:
    if detail.empty:
        return pd.DataFrame(columns=list(_PROFIT_MAP.keys()))

    grp = (
        detail.groupby("client", as_index=False)
        .agg(
            billing_amount=("billing_amount", "sum"),
            cost          =("cost", "sum"),
            employees     =("employee_id", "nunique"),
        )
    )
    grp["profit"]     = grp["billing_amount"] - grp["cost"]
    grp["margin_pct"] = (
        grp["profit"] / grp["billing_amount"].replace(0, float("nan")) * 100
    ).round(1)
    return grp.sort_values("billing_amount", ascending=False)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def save_organized_reports(
    detail_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    base_output_dir: str,
    month_str: str,
) -> tuple[str, str, str]:
    """
    Save reports in an organized month-based directory.

    Creates
    -------
    {base_output_dir}/{month_str}/
      final.xlsx          – billing + employee detail + internal summary
      issues.xlsx         – data-quality issues
      profitability.xlsx  – profitability analysis (separate file)

    Returns
    -------
    (final_path, issues_path, profitability_path)
    """
    month_dir   = os.path.join(base_output_dir, month_str)
    os.makedirs(month_dir, exist_ok=True)

    final_path  = os.path.join(month_dir, "final.xlsx")
    issues_path = os.path.join(month_dir, "issues.xlsx")
    profit_path = os.path.join(month_dir, "profitability.xlsx")

    # ── final.xlsx — billing + summary + employee detail ─────────────────────
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "דוח חיוב לקוחות"
    billing_report_df = _build_billing_report(detail_df)
    _write_sheet(ws1, billing_report_df, _BILLING_REPORT_MAP,
                 'דוח חיוב חודשי לפני מע"מ', add_totals=True)
    _style_billing_report(ws1, billing_report_df)

    ws2 = wb.create_sheet("סיכום פנימי")
    _write_sheet(ws2, _build_summary(detail_df), _SUMMARY_MAP,
                 "סיכום חיוב חודשי לפי לקוח", add_totals=True)

    ws3 = wb.create_sheet("פירוט לפי עובד")
    _write_sheet(ws3, detail_df, _DETAIL_MAP,
                 "פירוט לפי עובד ואתר", add_totals=True)

    wb.save(final_path)

    # ── profitability.xlsx ────────────────────────────────────────────────────
    wb_p    = Workbook()
    ws_prof = wb_p.active
    ws_prof.title = "ניתוח רווחיות"
    _write_sheet(ws_prof, _build_profitability(detail_df), _PROFIT_MAP,
                 "ניתוח רווחיות לפי לקוח", add_totals=True)
    wb_p.save(profit_path)

    # ── issues.xlsx ───────────────────────────────────────────────────────────
    wb2    = Workbook()
    ws_iss = wb2.active
    ws_iss.title = "חריגים"

    if not issues_df.empty:
        _write_sheet(ws_iss, issues_df, _ISSUES_MAP, "חריגים ובעיות שנמצאו")
        for row_idx in range(4, 4 + len(issues_df)):
            itype = ws_iss.cell(row_idx, 4).value or ""
            fill  = _ERROR_FILL if "חסר" in str(itype) or "שגיאה" in str(itype) \
                    else _WARNING_FILL
            for col_idx in range(1, len(_ISSUES_MAP) + 1):
                ws_iss.cell(row_idx, col_idx).fill = fill
    else:
        ws_iss.cell(1, 1).value = "✓ לא נמצאו חריגים"
        ws_iss.cell(1, 1).font  = Font(bold=True, color="008000", size=12)
        ws_iss.sheet_view.rightToLeft = True

    wb2.save(issues_path)

    return final_path, issues_path, profit_path


def save_reports(
    detail_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    output_dir: str,
) -> tuple[str, str]:
    """
    Legacy timestamped save (kept for backward compatibility with direct CLI usage).

    Returns
    -------
    (billing_path, issues_path)
    """
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    billing_path = os.path.join(output_dir, f"final_{ts}.xlsx")
    issues_path  = os.path.join(output_dir, f"issues_{ts}.xlsx")

    # ---- billing workbook ----
    wb = Workbook()

    # Sheet 1 – client billing report (main, matches reference format)
    ws1 = wb.active
    ws1.title = "דוח חיוב לקוחות"
    billing_report_df = _build_billing_report(detail_df)
    _write_sheet(ws1, billing_report_df, _BILLING_REPORT_MAP,
                 'דוח חיוב חודשי לפני מע"מ', add_totals=True)
    _style_billing_report(ws1, billing_report_df)

    # Sheet 2 – internal summary (with profitability)
    ws2 = wb.create_sheet("סיכום פנימי")
    _write_sheet(ws2, _build_summary(detail_df), _SUMMARY_MAP,
                 "סיכום חיוב חודשי לפי לקוח", add_totals=True)

    # Sheet 3 – employee detail
    ws3 = wb.create_sheet("פירוט לפי עובד")
    _write_sheet(ws3, detail_df, _DETAIL_MAP,
                 "פירוט לפי עובד ואתר", add_totals=True)

    # Sheet 4 – profitability
    ws4 = wb.create_sheet("ניתוח רווחיות")
    _write_sheet(ws4, _build_profitability(detail_df), _PROFIT_MAP,
                 "ניתוח רווחיות לפי לקוח", add_totals=True)

    wb.save(billing_path)

    # ---- issues workbook ----
    wb2 = Workbook()
    ws_iss = wb2.active
    ws_iss.title = "חריגים"

    if not issues_df.empty:
        # Color-code rows by issue severity
        _write_sheet(ws_iss, issues_df, _ISSUES_MAP, "חריגים ובעיות שנמצאו")
        # Highlight error rows in red, warning in yellow
        for row_idx in range(4, 4 + len(issues_df)):
            itype = ws_iss.cell(row_idx, 4).value or ""
            fill  = _ERROR_FILL if "חסר" in str(itype) or "שגיאה" in str(itype) else _WARNING_FILL
            for col_idx in range(1, len(_ISSUES_MAP) + 1):
                ws_iss.cell(row_idx, col_idx).fill = fill
    else:
        ws_iss.cell(1, 1).value = "✓ לא נמצאו חריגים"
        ws_iss.cell(1, 1).font  = Font(bold=True, color="008000", size=12)
        ws_iss.sheet_view.rightToLeft = True

    wb2.save(issues_path)

    return billing_path, issues_path


# ---------------------------------------------------------------------------
# Flat clean export (no merged headers, one row per employee per day)
# ---------------------------------------------------------------------------

def save_clean_export(detail_daily_df: pd.DataFrame, output_path: str) -> str:
    """
    Write a flat Excel with no merged headers and no extra rows.

    Columns (order guaranteed):
      client | site | employee_id | date | hours | completion_hours | rate | billing_amount

    One row per employee per worked day.  Suitable for import into external tools.
    """
    col_map = {
        "client":         "client",
        "site":           "site",
        "employee_id":    "employee_id",
        "date":           "date",
        "hours_to_pay":   "hours",
        "completion_day": "completion_hours",
        "rate":           "rate",
        "billing_day":    "billing_amount",
    }
    avail  = {k: v for k, v in col_map.items() if k in detail_daily_df.columns}
    out_df = (
        detail_daily_df[list(avail.keys())]
        .rename(columns=avail)
        .sort_values(
            [c for c in ("client", "employee_id", "date") if c in avail.values()]
        )
        .reset_index(drop=True)
    )
    out_df.to_excel(output_path, index=False)
    return output_path
