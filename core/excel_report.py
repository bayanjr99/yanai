"""
core/excel_report.py — Production Excel report from the merged cost DataFrame.

Usage:
    from core.excel_report import export_cost_report
    path = export_cost_report(df, output_path="cost_analysis_report.xlsx")

The function accepts the combined DataFrame returned by merge_and_allocate()
(one row per employee × site × month) and writes a fully-formatted, multi-sheet
Excel workbook.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# ── Style constants ───────────────────────────────────────────────────────────

_BLUE_DARK   = "1F4E79"
_BLUE_MED    = "2E75B6"
_BLUE_LIGHT  = "D6E4F0"
_STRIPE      = "EBF3FB"
_TOTAL_FILL  = "BDD7EE"
_WHITE       = "FFFFFF"
_ALERT_RED   = "FFC7CE"
_ALERT_FONT  = "9C0006"
_GREEN       = "C6EFCE"
_GREEN_FONT  = "276221"

_THIN  = Side(style="thin",   color="B0C4DE")
_THICK = Side(style="medium", color=_BLUE_DARK)

_BORDER       = Border(left=_THIN,  right=_THIN,  top=_THIN,  bottom=_THIN)
_BORDER_THICK = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)

_HEADER_FILL = PatternFill("solid", fgColor=_BLUE_DARK)
_TITLE_FILL  = PatternFill("solid", fgColor=_BLUE_MED)
_TOTAL_FILL_ = PatternFill("solid", fgColor=_TOTAL_FILL)
_STRIPE_FILL = PatternFill("solid", fgColor=_STRIPE)
_RED_FILL    = PatternFill("solid", fgColor=_ALERT_RED)
_GREEN_FILL  = PatternFill("solid", fgColor=_GREEN)

_HEADER_FONT = Font(bold=True,  color=_WHITE, size=10, name="Calibri")
_TITLE_FONT  = Font(bold=True,  color=_WHITE, size=12, name="Calibri")
_TOTAL_FONT  = Font(bold=True,  size=10, name="Calibri")
_DATA_FONT   = Font(size=10,    name="Calibri")
_ALERT_FONT_ = Font(bold=True,  color=_ALERT_FONT, size=10, name="Calibri")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_RIGHT  = Alignment(horizontal="right",  vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")
_WRAP   = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ILS0  = '₪#,##0'
_ILS2  = '₪#,##0.00'
_HRS2  = '#,##0.00'
_INT   = '#,##0'
_PCT   = '0.0%'


# ── Low-level cell writer ─────────────────────────────────────────────────────

def _cell(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    *,
    fmt: str | None = None,
    font: Font | None = None,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
    border: Border | None = _BORDER,
):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:    c.number_format = fmt
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c


def _auto_width(ws: Worksheet, min_w: int = 8, max_w: int = 42) -> None:
    for col_cells in ws.columns:
        width = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            min(max_w, max(min_w, width + 2))
        )


def _title_row(ws: Worksheet, text: str, n_cols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    _cell(ws, row, 1, text,
          font=_TITLE_FONT, fill=_TITLE_FILL, align=_CENTER, border=None)
    ws.row_dimensions[row].height = 22


def _header_row(ws: Worksheet, headers: list[str], row: int = 2) -> None:
    for ci, h in enumerate(headers, 1):
        _cell(ws, row, ci, h,
              font=_HEADER_FONT, fill=_HEADER_FILL, align=_WRAP)
    ws.row_dimensions[row].height = 28


def _total_row(
    ws: Worksheet,
    row: int,
    col_defs: list[tuple],
    totals: dict[int, float],
    label_col: int = 1,
    label: str = 'סה"כ',
) -> None:
    for ci, (_, fmt, _is_sum) in enumerate(col_defs, 1):
        val = totals.get(ci) if _is_sum else None
        if ci == label_col and val is None:
            val = label
        _cell(ws, row, ci, val,
              fmt=fmt, font=_TOTAL_FONT, fill=_TOTAL_FILL_, align=_RIGHT,
              border=_BORDER_THICK)
    ws.row_dimensions[row].height = 18


def _data_rows(
    ws: Worksheet,
    df: pd.DataFrame,
    col_defs: list[tuple],  # [(col_name, fmt, is_summed), ...]
    start_row: int = 3,
) -> tuple[int, dict[int, float]]:
    """
    Write data rows. Returns (last_row_written, totals_dict).
    col_defs entries:  (df_col_or_None, number_fmt, include_in_total)
    """
    totals: dict[int, float] = {}
    for ri, (_, row_data) in enumerate(df.iterrows()):
        r = start_row + ri
        stripe = _STRIPE_FILL if ri % 2 == 0 else None
        for ci, (col_name, fmt, is_sum) in enumerate(col_defs, 1):
            val = row_data.get(col_name) if col_name is not None else None
            if isinstance(val, float) and pd.isna(val):
                val = None
            _cell(ws, r, ci, val,
                  fmt=fmt, font=_DATA_FONT, fill=stripe,
                  align=_RIGHT if isinstance(val, (int, float)) else _LEFT)
            if is_sum and isinstance(val, (int, float)) and val is not None:
                totals[ci] = totals.get(ci, 0.0) + val
        ws.row_dimensions[r].height = 16
    return start_row + len(df) - 1, totals


def _month_sort_key(m: str) -> tuple:
    try:
        mm, yy = str(m).split("-")
        return int(yy), int(mm)
    except Exception:
        return 9999, 99


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_detail(wb: Workbook, df: pd.DataFrame, title: str) -> None:
    ws = wb.create_sheet("פירוט מלא")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A3"

    # Choose columns present in df
    pct_cols = [(f"h{p}", f"{p}%") for p in ["100", "125", "150", "175", "200"]
                if f"h{p}" in df.columns]

    headers = [
        "חודש", "מס' עובד", "שם עובד", "מדינה", "לקוח", "אתר",
        'סה"כ שעות',
    ] + [lbl for _, lbl in pct_cols] + [
        "עלות/שעה (₪)", "עלות מוקצה (₪)", "עלות מעביד (₪)",
    ]
    if "adjusted_levy" in df.columns:
        headers.append("אגרות מתוקנות (₪)")

    col_defs = [
        ("month",          None,   False),
        ("employee_id",    None,   False),
        ("employee_name",  None,   False),
        ("country",        None,   False),
        ("client",         None,   False),
        ("site",           None,   False),
        ("total_hours",    _HRS2,  True),
    ] + [(col, _HRS2, True) for col, _ in pct_cols] + [
        ("cost_per_hour",  _ILS2,  False),
        ("allocated_cost", _ILS0,  True),
        ("employer_cost",  _ILS0,  True),
    ]
    if "adjusted_levy" in df.columns:
        col_defs.append(("adjusted_levy", _ILS0, True))

    sorted_df = df.sort_values(
        ["month", "employee_id", "client"],
        key=lambda s: s.map(_month_sort_key) if s.name == "month" else s,
    ).reset_index(drop=True)

    _title_row(ws, title, len(headers))
    _header_row(ws, headers)
    last_r, totals = _data_rows(ws, sorted_df, col_defs)
    _total_row(ws, last_r + 1, col_defs, totals, label_col=1)
    _auto_width(ws)


def _sheet_summary(
    wb: Workbook,
    df: pd.DataFrame,
    group_col: str,
    sheet_name: str,
    group_label: str,
    title: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A3"

    agg = (
        df[df[group_col].astype(str).str.strip() != ""]
        .groupby(group_col, as_index=False)
        .agg(
            total_hours   =("total_hours",    "sum"),
            total_cost    =("allocated_cost", "sum"),
            employer_cost =("employer_cost",  "sum"),
        )
    )
    _safe = agg["total_hours"].replace(0, float("nan"))
    agg["avg_cph"] = (agg["total_cost"] / _safe).round(2)
    agg = agg.sort_values("total_cost", ascending=False).reset_index(drop=True)

    headers  = [group_label, 'סה"כ שעות', 'סה"כ עלות (₪)', "עלות ממוצעת/שעה (₪)"]
    col_defs = [
        (group_col,      None,   False),
        ("total_hours",  _HRS2,  True),
        ("total_cost",   _ILS0,  True),
        ("avg_cph",      _ILS2,  False),
    ]

    _title_row(ws, title, len(headers))
    _header_row(ws, headers)
    last_r, totals = _data_rows(ws, agg, col_defs)
    _total_row(ws, last_r + 1, col_defs, totals, label_col=1)
    _auto_width(ws)


def _sheet_employees(wb: Workbook, df: pd.DataFrame, title: str) -> None:
    ws = wb.create_sheet("לפי עובד")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A3"

    agg = (
        df.groupby(["employee_id", "employee_name"], as_index=False)
        .agg(
            country      =("country",        "first"),
            total_hours  =("total_hours",    "sum"),
            employer_cost=("employer_cost",  "first"),  # per-employee, not per-site
        )
    )
    _safe = agg["total_hours"].replace(0, float("nan"))
    agg["avg_cph"] = (agg["employer_cost"] / _safe).round(2)
    agg = agg.sort_values("employer_cost", ascending=False).reset_index(drop=True)

    headers  = ["מס' עובד", "שם עובד", "מדינה", 'סה"כ שעות', "עלות מעביד (₪)", "עלות ממוצעת/שעה (₪)"]
    col_defs = [
        ("employee_id",   None,   False),
        ("employee_name", None,   False),
        ("country",       None,   False),
        ("total_hours",   _HRS2,  True),
        ("employer_cost", _ILS0,  True),
        ("avg_cph",       _ILS2,  False),
    ]

    _title_row(ws, title, len(headers))
    _header_row(ws, headers)
    last_r, totals = _data_rows(ws, agg, col_defs)
    _total_row(ws, last_r + 1, col_defs, totals, label_col=1)
    _auto_width(ws)


def _sheet_country(wb: Workbook, df: pd.DataFrame, title: str) -> None:
    ws = wb.create_sheet("לפי מדינה")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A3"

    country_df = df[df["country"].astype(str).str.strip() != ""].copy()
    if country_df.empty:
        ws.cell(row=1, column=1, value="אין נתוני מדינה")
        return

    agg = (
        country_df.groupby("country", as_index=False)
        .agg(
            employees    =("employee_id",   "nunique"),
            total_hours  =("total_hours",   "sum"),
            total_cost   =("allocated_cost","sum"),
        )
    )
    _safe = agg["total_hours"].replace(0, float("nan"))
    agg["avg_cph"]    = (agg["total_cost"] / _safe).round(2)
    agg["total_hours"] = agg["total_hours"].round(2)
    agg["total_cost"]  = agg["total_cost"].round(2)
    agg = agg.sort_values("total_cost", ascending=False).reset_index(drop=True)

    headers  = ["מדינה", "עובדים", 'סה"כ שעות', 'סה"כ עלות (₪)', "עלות ממוצעת/שעה (₪)"]
    col_defs = [
        ("country",     None,   False),
        ("employees",   _INT,   False),
        ("total_hours", _HRS2,  True),
        ("total_cost",  _ILS0,  True),
        ("avg_cph",     _ILS2,  False),
    ]

    _title_row(ws, title, len(headers))
    _header_row(ws, headers)
    last_r, totals = _data_rows(ws, agg, col_defs)
    _total_row(ws, last_r + 1, col_defs, totals, label_col=1)
    _auto_width(ws)

    # Secondary table: Country × Client cost breakdown
    cc_row_start = last_r + 4
    ws.cell(row=cc_row_start, column=1, value="התפלגות מדינה × לקוח (₪)").font = _TOTAL_FONT
    pivot = (
        country_df.groupby(["country", "client"])["allocated_cost"]
        .sum()
        .unstack(fill_value=0)
        .round(0)
    )
    pivot["סה\"כ"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("סה\"כ", ascending=False)

    pivot_df = pivot.reset_index()
    for ci, col in enumerate(pivot_df.columns, 1):
        _cell(ws, cc_row_start + 1, ci, str(col),
              font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    for ri, (_, row) in enumerate(pivot_df.iterrows(), 2):
        stripe = _STRIPE_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            fmt = _ILS0 if ci > 1 else None
            _cell(ws, cc_row_start + ri, ci, val,
                  fmt=fmt, font=_DATA_FONT, fill=stripe,
                  align=_RIGHT if ci > 1 else _LEFT)
    _auto_width(ws)


def _sheet_alerts(
    wb: Workbook,
    df: pd.DataFrame,
    title: str,
    threshold: float = 200.0,
) -> None:
    ws = wb.create_sheet("התראות")
    ws.sheet_view.rightToLeft = True

    current_row = 1
    _title_row(ws, title, 5, row=current_row)
    current_row += 2

    def _section(label: str, rows: list[dict], columns: list[str], headers: list[str]) -> None:
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=f"⚠ {label}").font = Font(bold=True, size=11, name="Calibri")
        current_row += 1

        if not rows:
            c = ws.cell(row=current_row, column=1, value="✓ לא נמצאו חריגים")
            c.font = Font(color=_GREEN_FONT, bold=True, name="Calibri")
            c.fill = _GREEN_FILL
            current_row += 2
            return

        for ci, h in enumerate(headers, 1):
            _cell(ws, current_row, ci, h,
                  font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
        current_row += 1

        for ri, row in enumerate(rows):
            stripe = _RED_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFE0E0")
            for ci, col in enumerate(columns, 1):
                val = row.get(col)
                _cell(ws, current_row, ci, val,
                      font=_ALERT_FONT_, fill=stripe,
                      align=_RIGHT if isinstance(val, (int, float)) else _LEFT)
            current_row += 1
        current_row += 1  # blank separator

    # ── 1. Zero hours ────────────────────────────────────────────────────────
    zero_h = (
        df[df["total_hours"] == 0]
        .drop_duplicates("employee_id")[["employee_id", "employee_name", "country", "month"]]
        .rename(columns={"month": "חודש"})
    )
    _section(
        "עובדים עם 0 שעות",
        zero_h.to_dict("records"),
        ["employee_id", "employee_name", "country", "חודש"],
        ["מס' עובד", "שם עובד", "מדינה", "חודש"],
    )

    # ── 2. Cost but no hours ─────────────────────────────────────────────────
    cost_no_hours = (
        df[(df["employer_cost"] > 0) & (df["total_hours"] == 0)]
        .drop_duplicates("employee_id")
        [["employee_id", "employee_name", "country", "employer_cost", "month"]]
    )
    _section(
        "עלות ללא שעות",
        cost_no_hours.to_dict("records"),
        ["employee_id", "employee_name", "country", "employer_cost", "month"],
        ["מס' עובד", "שם עובד", "מדינה", "עלות (₪)", "חודש"],
    )

    # ── 3. High cost per hour ────────────────────────────────────────────────
    high_cph = (
        df[df["cost_per_hour"] > threshold]
        .drop_duplicates("employee_id")
        .sort_values("cost_per_hour", ascending=False)
        [["employee_id", "employee_name", "country", "cost_per_hour", "employer_cost"]]
    )
    _section(
        f"עלות לשעה גבוהה (מעל ₪{threshold:,.0f})",
        high_cph.to_dict("records"),
        ["employee_id", "employee_name", "country", "cost_per_hour", "employer_cost"],
        ["מס' עובד", "שם עובד", "מדינה", "עלות/שעה (₪)", "עלות מעביד (₪)"],
    )

    # ── 4. Missing client or site ────────────────────────────────────────────
    missing = df[
        df["client"].astype(str).str.strip().isin(["", "nan", "None"]) |
        df["site"].astype(str).str.strip().isin(["", "nan", "None"])
    ].drop_duplicates("employee_id")[["employee_id", "employee_name", "client", "site", "month"]]
    _section(
        "חסרים לקוח או אתר",
        missing.to_dict("records"),
        ["employee_id", "employee_name", "client", "site", "month"],
        ["מס' עובד", "שם עובד", "לקוח", "אתר", "חודש"],
    )

    # ── 5. Missing country ───────────────────────────────────────────────────
    no_country = df[
        df["country"].astype(str).str.strip().isin(["", "nan", "None"])
    ].drop_duplicates("employee_id")[["employee_id", "employee_name", "month"]]
    _section(
        "חסרת מדינה",
        no_country.to_dict("records"),
        ["employee_id", "employee_name", "month"],
        ["מס' עובד", "שם עובד", "חודש"],
    )

    # ── 6. Allocation mismatch ───────────────────────────────────────────────
    alloc = (
        df.groupby("employee_id", as_index=False)
        .agg(sum_alloc=("allocated_cost", "sum"),
             emp_cost  =("employer_cost",  "first"))
    )
    alloc["diff"] = (alloc["sum_alloc"] - alloc["emp_cost"]).abs()
    mismatch = alloc[alloc["diff"] > 1.0].merge(
        df[["employee_id", "employee_name"]].drop_duplicates(), on="employee_id"
    )
    _section(
        "אי-התאמה בהקצאת עלות",
        mismatch.to_dict("records"),
        ["employee_id", "employee_name", "sum_alloc", "emp_cost", "diff"],
        ["מס' עובד", "שם עובד", "סכום מוקצה (₪)", "עלות מעביד (₪)", "הפרש (₪)"],
    )

    _auto_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def export_cost_report(
    df: pd.DataFrame,
    output_path: str = "cost_analysis_report.xlsx",
    threshold: float = 200.0,
    months_label: str = "",
) -> str:
    """
    Generate a complete, formatted Excel cost report from the merged DataFrame.

    Parameters
    ----------
    df            Combined DataFrame from merge_and_allocate() (all months).
    output_path   Destination file path.
    threshold     cost_per_hour alert threshold (default ₪200/h).
    months_label  Optional subtitle string, e.g. "01-2025 עד 03-2026".

    Returns the absolute path of the written file.
    """
    if df is None or df.empty:
        raise ValueError("DataFrame is empty — nothing to export.")

    # Normalise key columns
    df = df.copy()
    for col in ("client", "site", "country", "employee_name", "employee_id", "month"):
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    for col in ("total_hours", "allocated_cost", "employer_cost", "cost_per_hour"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    months_range = months_label or (
        " – ".join([
            min(df["month"], key=_month_sort_key),
            max(df["month"], key=_month_sort_key),
        ])
        if not df.empty else ""
    )
    title = f"דוח עלויות עובדים — {months_range}" if months_range else "דוח עלויות עובדים"

    wb = Workbook()
    wb.remove(wb.active)   # drop default blank sheet

    _sheet_detail(wb, df, title)
    _sheet_summary(wb, df, "client", "לפי לקוח",   "לקוח", title)
    _sheet_summary(wb, df, "site",   "לפי אתר",    "אתר",  title)
    _sheet_employees(wb, df, title)
    _sheet_country(wb, df, title)
    _sheet_alerts(wb, df, title, threshold=threshold)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out.resolve())
