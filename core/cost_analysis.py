"""
core/cost_analysis.py — Employee cost analysis by employee_id.

Key rule: ALL joins use employee_id (string). Never use employee_name.

Flow:
  1. load_hours_from_pdf()   → aggregate daily PDF rows → monthly hours per employee × site
  2. load_costs_xlsx()       → employer cost per employee from costs.xlsx
  3. merge_and_allocate()    → join on employee_id, split cost by hours ratio
  4. build_sheets()          → create the 4 output DataFrames
  5. detect_warnings()       → data quality issues
  6. export_to_excel()       → write cost_analysis.xlsx
"""

from __future__ import annotations

import re
import warnings as _warnings
from pathlib import Path

import pandas as pd


# ---------------------------------------------------------------------------
# 1.  Hours from PDF
# ---------------------------------------------------------------------------

def load_hours_from_pdf(pdf_path: str, month: str) -> pd.DataFrame:
    """
    Parse an Andromeda payroll PDF and return one row per (employee_id, site)
    aggregated for the whole month.

    Uses the existing pdf_parser — no name matching, only employee_id.

    Returns columns:
      month, employee_id, employee_name, site,
      work_days, total_hours, break_hours
    """
    from core.pdf_parser import parse_pdf

    daily = parse_pdf(pdf_path)
    if daily.empty:
        return pd.DataFrame(columns=[
            "month", "employee_id", "employee_name", "site",
            "work_days", "total_hours", "break_hours",
        ])

    # Normalize employee_id to string (key column — never touch name)
    daily["employee_id"] = daily["employee_id"].astype(str).str.strip()

    # Aggregate: one row per employee × site
    agg = (
        daily
        .groupby(["employee_id", "employee_name", "site"], as_index=False)
        .agg(
            work_days  =("date",         "nunique"),
            total_hours=("hours_to_pay", "sum"),
            break_hours=("break_hours",  "sum"),
        )
    )

    agg["total_hours"] = agg["total_hours"].round(2)
    agg["break_hours"] = agg["break_hours"].round(2)
    agg["month"]       = month

    cols = ["month", "employee_id", "employee_name", "site",
            "work_days", "total_hours", "break_hours"]
    return agg[cols].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 1b. Hours from Excel
# ---------------------------------------------------------------------------

def load_hours_from_xlsx(xlsx_path: str, month: str) -> pd.DataFrame:
    """
    Parse an hours.xlsx file and return one row per (employee_id, site)
    aggregated for the whole month.  Supports both daily-breakdown files
    (with a date column) and already-aggregated monthly files.

    Returns same schema as load_hours_from_pdf:
      month, employee_id, employee_name, site, work_days, total_hours, break_hours
    """
    EMPTY = pd.DataFrame(columns=[
        "month", "employee_id", "employee_name", "site",
        "work_days", "total_hours", "break_hours",
    ])

    try:
        df = pd.read_excel(xlsx_path, dtype=str)
    except Exception:
        return EMPTY

    if df.empty:
        return EMPTY

    def _find_col(hints: list[str]) -> str | None:
        for col in df.columns:
            col_n = str(col).lower().replace(" ", "").replace('"', "")
            for h in hints:
                if h.lower().replace(" ", "").replace('"', "") in col_n:
                    return col
        return None

    emp_id_col   = _find_col(["מספרעובד", "מסעובד", "employee_id", "id", "מס'עובד"])
    emp_name_col = _find_col(["שםעובד", "עובד", "employee_name", "שם"])
    site_col     = _find_col(["שםפרויקט", "פרויקט", "אתר", "site", "locality"])
    date_col     = _find_col(["תאריך", "date", "יום"])
    hours_col    = _find_col(['סהכשעות', "total_hours", "שעות", "שעותעבודה", "hours"])

    if emp_id_col is None or hours_col is None:
        return EMPTY

    df["_emp_id"] = df[emp_id_col].astype(str).str.strip()
    df = df[df["_emp_id"].str.fullmatch(r"\d{3,6}")].copy()
    if df.empty:
        return EMPTY

    df["_emp_name"] = (df[emp_name_col].astype(str).str.strip()
                       if emp_name_col else "")
    df["_site"]     = (df[site_col].astype(str).str.strip()
                       if site_col else "")
    df["_hours"]    = pd.to_numeric(df[hours_col], errors="coerce").fillna(0.0)

    if date_col is not None:
        # Daily file: group by (employee_id, employee_name, site)
        df["_date"] = df[date_col].astype(str).str.strip()
        agg = (
            df.groupby(["_emp_id", "_emp_name", "_site"], as_index=False)
            .agg(
                work_days  =("_date",   "nunique"),
                total_hours=("_hours",  "sum"),
            )
        )
    else:
        # Already-aggregated file: each row is already per (employee × site)
        days_col = _find_col(["ימיעבודה", "ימים", "days"])
        df["_days"] = (pd.to_numeric(df[days_col], errors="coerce").fillna(0.0)
                       if days_col else 0.0)
        agg = (
            df.groupby(["_emp_id", "_emp_name", "_site"], as_index=False)
            .agg(
                work_days  =("_days",   "sum"),
                total_hours=("_hours",  "sum"),
            )
        )

    agg = agg.rename(columns={
        "_emp_id":   "employee_id",
        "_emp_name": "employee_name",
        "_site":     "site",
    })
    agg["total_hours"] = agg["total_hours"].round(2)
    agg["break_hours"] = 0.0
    agg["month"]       = month

    cols = ["month", "employee_id", "employee_name", "site",
            "work_days", "total_hours", "break_hours"]
    return agg[cols].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 1c. Worked-days from daily data (for accurate levy)
# ---------------------------------------------------------------------------

def compute_month_working_days(hours_path: str) -> int | None:
    """
    Return the total number of unique working dates in the month's source file.

    This is the correct denominator for levy proration:
        adjusted_levy = full_monthly_levy × employee_work_days / month_work_days

    Returns None when the source has no daily date breakdown (already-aggregated Excel).
    """
    ext = hours_path.lower()
    if ext.endswith(".xls") and not ext.endswith(".xlsx"):
        ext = ext[:-4] + ".xlsx"

    if ext.endswith(".pdf"):
        from core.pdf_parser import parse_pdf
        try:
            daily = parse_pdf(hours_path)
        except (FileNotFoundError, OSError):
            return None
        if daily.empty or "date" not in daily.columns:
            return None
        n = int(daily["date"].nunique())
        return n if n > 0 else None

    if ext.endswith(".xlsx"):
        try:
            df = pd.read_excel(hours_path, dtype=str)
        except Exception:
            return None

        def _fc(hints: list[str]) -> str | None:
            for col in df.columns:
                cn = str(col).lower().replace(" ", "").replace('"', "")
                if any(h.lower().replace(" ", "") in cn for h in hints):
                    return col
            return None

        date_col = _fc(["תאריך", "date", "יום"])
        hrs_col  = _fc(['סהכשעות', "total_hours", "שעות", "שעותעבודה", "hours"])
        if date_col is None or hrs_col is None:
            return None
        df["_hours"] = pd.to_numeric(df[hrs_col], errors="coerce").fillna(0.0)
        worked = df[df["_hours"] > 0]
        n = int(worked[date_col].nunique())
        return n if n > 0 else None

    return None


def compute_worked_days(hours_path: str) -> "pd.Series | None":
    """
    Count unique working dates per employee from daily source data.

    Business rule:
      Any date where an employee has hours > 0 counts as ONE full working day.
      This is used for levy calculation (levy is per day, not per hour).

    Returns:
      pd.Series  indexed by employee_id (str), values = float (day count)
      None       when daily data is unavailable (already-aggregated Excel)
    """
    ext = hours_path.lower()

    # Treat .xls identically to .xlsx
    if ext.endswith(".xls") and not ext.endswith(".xlsx"):
        ext = ext[:-4] + ".xlsx"

    if ext.endswith(".pdf"):
        from core.pdf_parser import parse_pdf
        daily = parse_pdf(hours_path)
        if daily.empty:
            return None
        daily["employee_id"] = daily["employee_id"].astype(str).str.strip()
        worked = daily[daily["hours_to_pay"] > 0]
        if worked.empty:
            return None
        return (
            worked.groupby("employee_id")["date"]
            .nunique()
            .rename("worked_days")
            .astype(float)
        )

    if ext.endswith(".xlsx"):
        try:
            df = pd.read_excel(hours_path, dtype=str)
        except Exception:
            return None

        def _fc(hints: list[str]) -> str | None:
            for col in df.columns:
                cn = str(col).lower().replace(" ", "").replace('"', "")
                if any(h.lower().replace(" ", "") in cn for h in hints):
                    return col
            return None

        emp_col  = _fc(["מספרעובד", "מסעובד", "employee_id", "id"])
        date_col = _fc(["תאריך", "date", "יום"])
        hrs_col  = _fc(['סהכשעות', "total_hours", "שעות", "שעותעבודה", "hours"])

        if emp_col is None or date_col is None or hrs_col is None:
            return None

        df["_emp_id"] = df[emp_col].astype(str).str.strip()
        df = df[df["_emp_id"].str.fullmatch(r"\d{3,6}")].copy()
        df["_hours"] = pd.to_numeric(df[hrs_col], errors="coerce").fillna(0.0)
        worked = df[df["_hours"] > 0]
        if worked.empty:
            return None
        return (
            worked.groupby("_emp_id")[date_col]
            .nunique()
            .rename("worked_days")
            .rename_axis("employee_id")
            .astype(float)
        )

    return None


# ---------------------------------------------------------------------------
# 1f. Hours from Andromeda payroll-detail PDF (דוח פירוט שכר)
# ---------------------------------------------------------------------------

def _un_reverse_hebrew(s: str) -> str:
    """
    pdfplumber extracts RTL Hebrew text in visual order: each individual
    word's characters are reversed, and the word order is also reversed.

    Example: "יללכ-המקש תיב" → "בית שקמה-כללי"

    Numbers, ASCII, and mixed tokens are left unchanged (they're in correct
    order in the PDF byte stream already).
    """
    HEB = re.compile(r'[א-תװ-״יִ-פֿ"\'"\-]+')

    def _rev_word(w: str) -> str:
        if HEB.search(w):
            return w[::-1]
        return w

    words = s.split()
    return " ".join(_rev_word(w) for w in reversed(words))


def _extract_all_site_clients(text: str) -> list[tuple[str, str]]:
    """
    Parse ALL (client, site) pairs from an Andromeda payroll-PDF page text.

    pdfplumber extracts RTL Hebrew text in visual order: each word's
    characters are reversed, and word order is left-to-right (reversed).

    Section-header lines have the form (visual LTR):
      [site_name_rev] "רתא םש" [client_name_rev] "חוקל םש"

    Where "רתא םש" = visual of "שם אתר" and "חוקל םש" = visual of "שם לקוח".

    Returns list of (client, site) tuples in the order they appear on the page.
    One tuple per table group; the i-th tuple maps to the i-th table.
    """
    SITE_MARKER   = "רתא םש"    # reversed "שם אתר"
    CLIENT_MARKER = "חוקל םש"   # reversed "שם לקוח"

    results: list[tuple[str, str]] = []

    for line in text.splitlines():
        if SITE_MARKER not in line:
            continue
        # Line format: [site_rev] SITE_MARKER [client_rev] CLIENT_MARKER
        site_parts = line.split(SITE_MARKER, 1)
        site_rev   = site_parts[0].strip()
        rest       = site_parts[1].strip() if len(site_parts) > 1 else ""

        # Remove CLIENT_MARKER suffix to isolate client_rev
        if CLIENT_MARKER in rest:
            client_rev = rest.split(CLIENT_MARKER)[0].strip()
        else:
            client_rev = rest.strip()

        client = _un_reverse_hebrew(client_rev)
        site   = _un_reverse_hebrew(site_rev)
        results.append((client, site))

    return results


def load_hours_from_payroll_pdf(pdf_path: str, month: str) -> pd.DataFrame:
    """
    Parse an Andromeda payroll-detail PDF (דוח פירוט שכר).

    This is the PRIMARY hours source.  The PDF is organised as a series of
    client+site groups; each group has a column-header row followed by data
    rows (one per employee) and a subtotal row.

    Column layout (consistent across all months, 13 cols, RTL order):
      col  6  תועש 150%       → h150
      col  7  תועש 125%       → h125
      col  8  תועש 100%       → h100
      col  9  סמ ימי הדובע    → work_days
      col 10  רפסמ ןוכרד      → passport (used for subtotal detection)
      col 11  דבוע םש         → employee_name
      col 12  רפסמ דבוע       → employee_id

    Client and site are extracted from section-header lines in the page text.

    Returns columns:
      month, employee_id, employee_name, client, site,
      work_days, total_hours, break_hours, h100, h125, h150
    """
    EMPTY = pd.DataFrame(columns=[
        "month", "employee_id", "employee_name", "client", "site",
        "work_days", "total_hours", "break_hours", "h100", "h125", "h150",
    ])

    try:
        import pdfplumber as _pdfp
    except ImportError:
        return EMPTY

    EMP_ID_RE    = re.compile(r"^\d{3,6}$")
    PASSPORT_RE  = re.compile(r"^[A-Z]{1,2}\d{5,}$")

    def _n(v) -> float:
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", "").strip())
        except ValueError:
            return 0.0

    rows: list[dict] = []

    with _pdfp.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""

            # Collect all (client, site) pairs for this page in top-to-bottom order.
            # Each section-header row in a table triggers the NEXT context pair.
            # ctx_idx starts at 0; the header row consumes contexts[ctx_idx] for the
            # rows that follow — there is no pre-seeded initial context.
            contexts    = _extract_all_site_clients(page_text)
            ctx_idx     = 0
            current_client, current_site = ("", "")

            for tbl in page.extract_tables():
                if not tbl:
                    continue

                for row in tbl:
                    if not row:
                        continue

                    # ── Header row → set context for the rows that follow ───
                    if any(
                        v is not None and "תועש" in str(v)
                        for v in row[min(6, len(row)-1):min(9, len(row))]
                    ):
                        if ctx_idx < len(contexts):
                            current_client, current_site = contexts[ctx_idx]
                            ctx_idx += 1
                        continue

                    if len(row) < 13:
                        continue

                    emp_id_raw = str(row[12] or "").strip()
                    if not EMP_ID_RE.fullmatch(emp_id_raw):
                        continue   # subtotal, blank, or stray row

                    # ── Data row ────────────────────────────────────────────
                    h100      = _n(row[8])
                    h125      = _n(row[7])
                    h150      = _n(row[6])
                    work_days = _n(row[9])
                    name_raw  = str(row[11] or "").replace("\n", " ").strip()

                    rows.append({
                        "employee_id":   emp_id_raw,
                        "employee_name": name_raw,
                        "client":        current_client,
                        "site":          current_site,
                        "work_days":     work_days,
                        "h100":          h100,
                        "h125":          h125,
                        "h150":          h150,
                    })

    if not rows:
        return EMPTY

    df = pd.DataFrame(rows)
    df["employee_id"] = df["employee_id"].astype(str).str.strip()

    # One employee may appear in multiple site groups → keep all (multi-site)
    agg = (
        df.groupby(["employee_id", "employee_name", "client", "site"], as_index=False)
        .agg(
            work_days=("work_days", "sum"),
            h100     =("h100",      "sum"),
            h125     =("h125",      "sum"),
            h150     =("h150",      "sum"),
        )
    )

    agg["total_hours"] = (agg["h100"] + agg["h125"] + agg["h150"]).round(2)
    agg["break_hours"] = 0.0
    agg["month"]       = month

    cols = ["month", "employee_id", "employee_name", "client", "site",
            "work_days", "total_hours", "break_hours", "h100", "h125", "h150"]
    return agg[cols].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 1d. Hours from billing Excel (percentage-column format)
# ---------------------------------------------------------------------------

def load_hours_from_billing_xlsx(xlsx_path: str, month: str) -> pd.DataFrame:
    """
    Parse a billing Excel where each hour type is a separate column.

    Expected columns (detected by name, order-independent, Hebrew/English):
      - Employee ID  : מספר עובד / מס עובד / employee_id
      - Employee Name: שם עובד / employee_name
      - Client       : לקוח / customer / client
      - Site         : אתר / פרויקט / locality / site
      - Hour columns : 100% / 125% / 150% / 175% / 200%
                       (also accepted without %, e.g. "שעות 125")
      - Total hours  : סה"כ שעות / total_hours / hours
                       (used as fallback when no % columns are found)
      - Work days    : ימים / ימי עבודה / days (optional)

    Returns one row per (employee_id × client × site) with:
      month, employee_id, employee_name, client, site,
      work_days, total_hours, break_hours,
      h100, h125, h150, h175, h200   ← hours at each rate
    """
    EMPTY = pd.DataFrame(columns=[
        "month", "employee_id", "employee_name", "client", "site",
        "work_days", "total_hours", "break_hours",
    ])

    try:
        # Read bytes first: releases the OS file handle before pandas parses,
        # avoiding Windows exclusive-lock conflicts (e.g. Streamlit holding xlsx).
        import io as _io
        with open(xlsx_path, "rb") as _fh:
            _raw = _fh.read()
        df = pd.read_excel(_io.BytesIO(_raw), dtype=str)
    except Exception:
        return EMPTY

    if df.empty:
        return EMPTY

    def _norm(s: str) -> str:
        return s.lower().replace(" ", "").replace('"', "").replace("'", "").replace("%", "")

    def _find_col(hints: list[str]) -> str | None:
        # Iterate hints first so more-specific hints take priority over column order.
        for h in hints:
            hn = _norm(h)
            for col in df.columns:
                if hn in _norm(str(col)):
                    return col
        return None

    def _find_pct_col(pct: str) -> str | None:
        """
        Find the hours column for a specific rate percentage, e.g. '125%'.

        Two-pass strategy so exact names always beat partial matches:
          Pass 1 — exact / canonical names:
            "125%", "125", "125Hours", "h125"  (case-insensitive, no spaces)
          Pass 2 — starts with the number:
            column name starts with "125" followed by a non-digit
          Pass 3 — last resort broad match (number + % or Hours anywhere)
            Never matches columns where the number appears mid-sentence
            (e.g. "תעריף שעות 100 אחוז" which is a rate, not an hours column).
        """
        num = pct.rstrip("%")

        # Pass 1: exact or canonical names (normalised, no spaces)
        exact_forms = {num, pct, f"{num}hours", f"h{num}", f"hours{num}"}
        for col in df.columns:
            s_norm = str(col).strip().lower().replace(" ", "").replace("%", "")
            if s_norm in exact_forms:
                return col

        # Pass 2: column name STARTS with the number (e.g. "100 שעות", "100 %")
        for col in df.columns:
            s = str(col).strip()
            if re.match(rf"^{re.escape(num)}\D", s):
                return col

        return None

    emp_id_col   = _find_col(["מספרעובד", "מסעובד", "employee_id", "empid"])
    emp_name_col = _find_col(["שםעובד", "employee_name"])
    client_col   = _find_col(["לקוח", "customer", "client", "לקוחות"])
    site_col     = _find_col(["אתר", "פרויקט", "locality", "site", "project"])
    days_col     = _find_col(["ימיעבודה", "ימים", "days"])
    country_col  = _find_col(["מדינה", "country", "nationality", "לאום"])

    if emp_id_col is None:
        return EMPTY

    # ── Percentage-rate hour columns ──────────────────────────────────────────
    pct_rates = ["100%", "125%", "150%", "175%", "200%"]
    pct_map: dict[str, str | None] = {p: _find_pct_col(p) for p in pct_rates}

    # Debug: log column mapping so mismatches are easy to spot
    import logging as _log
    _logger = _log.getLogger(__name__)
    _logger.debug("load_hours_from_billing_xlsx column mapping for %s:", xlsx_path)
    _logger.debug("  employee_id  -> %r", emp_id_col)
    _logger.debug("  employee_name-> %r", emp_name_col)
    _logger.debug("  client       -> %r", client_col)
    _logger.debug("  site         -> %r", site_col)
    _logger.debug("  country      -> %r", country_col)
    _logger.debug("  work_days    -> %r", days_col)
    for pct, col in pct_map.items():
        sample = (
            str(df.iloc[0][col])[:12] if col is not None and not df.empty else "n/a"
        )
        _logger.debug("  %-5s        -> %r  (sample: %s)", pct, col, sample)

    for pct, col in pct_map.items():
        key = f"_h{pct.rstrip('%')}"
        df[key] = (
            pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            if col is not None else 0.0
        )

    found_pct_cols = [p for p, c in pct_map.items() if c is not None]

    if found_pct_cols:
        df["_total_hours"] = sum(df[f"_h{p.rstrip('%')}"] for p in found_pct_cols)
    else:
        # Fallback: use an explicit total-hours column
        total_col = _find_col(['סהכשעות', "total_hours", "שעות", "שעותעבודה", "hours"])
        if total_col is None:
            return EMPTY
        df["_total_hours"] = pd.to_numeric(df[total_col], errors="coerce").fillna(0.0)

    # ── Clean & filter rows ───────────────────────────────────────────────────
    df["_emp_id"] = df[emp_id_col].astype(str).str.strip()
    df = df[df["_emp_id"].str.fullmatch(r"\d{3,6}")].copy()
    if df.empty:
        return EMPTY

    df["_emp_name"] = df[emp_name_col].astype(str).str.strip() if emp_name_col else ""
    df["_client"]   = df[client_col].astype(str).str.strip()   if client_col   else ""
    df["_site"]     = df[site_col].astype(str).str.strip()     if site_col     else ""
    df["_country"]  = df[country_col].astype(str).str.strip()  if country_col  else ""
    df["_days"]     = (
        pd.to_numeric(df[days_col], errors="coerce").fillna(0.0) if days_col else 0.0
    )

    # ── Aggregate: one row per (employee × client × site) ────────────────────
    group_cols = ["_emp_id", "_emp_name", "_client", "_site", "_country"]
    pct_agg = {
        f"h{p.rstrip('%')}": (f"_h{p.rstrip('%')}", "sum")
        for p in pct_rates
    }

    agg = (
        df.groupby(group_cols, as_index=False)
        .agg(
            work_days   =("_days",        "sum"),
            total_hours =("_total_hours", "sum"),
            **pct_agg,
        )
    )
    agg = agg.rename(columns={
        "_emp_id":      "employee_id",
        "_emp_name":    "employee_name",
        "_client":      "client",
        "_site":        "site",
        "_country":     "country",
    })

    agg["total_hours"] = agg["total_hours"].round(2)
    agg["break_hours"] = 0.0
    agg["month"]       = month

    base_cols  = ["month", "employee_id", "employee_name", "client", "site", "country",
                  "work_days", "total_hours", "break_hours"]
    extra_cols = [f"h{p.rstrip('%')}" for p in pct_rates if f"h{p.rstrip('%')}" in agg.columns]

    return agg[base_cols + extra_cols].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 1e. Costs from PDF (Andromeda detailed cost report)
# ---------------------------------------------------------------------------

def load_costs_pdf(pdf_path: str) -> pd.DataFrame:
    """
    Parse an Andromeda 'דוח עלות מפורט' PDF (costs.pdf).

    Column layout (consistent across 14/15/16-col table variants):
      col 0  : סה"כ עלות  → employer_cost
      col 3  : פנסיה מעביד → pension
      col 5  : אגרות והיטלים → levy
      col 9  : ביטוח לאומי מעביד → bituach
      col 10 : משכורת ברוטו → gross_salary
      passport col  : detected dynamically (alphanumeric ID)
      employee_id   : column immediately after passport col

    One employee may appear in multiple rows (different site allocations).
    Their costs are summed, identical to how load_costs_xlsx() works.

    Returns same schema as load_costs_xlsx:
      employee_id, employer_cost, gross_salary, bituach, levy, pension, client, site
    """
    EMPTY = pd.DataFrame(columns=[
        "employee_id", "employer_cost", "gross_salary",
        "bituach", "levy", "pension", "client", "site",
    ])
    try:
        import pdfplumber
    except ImportError:
        return EMPTY

    PASSPORT_RE = re.compile(r"^[A-Z]{1,2}\d{5,}$")
    EMP_ID_RE   = re.compile(r"^\d{3,6}$")
    HEADER_KW   = {"תולע", "ירודיס", "דבוע סמ", "היסנפ", "וטורב"}

    def _n(v: object) -> float:
        if v is None or str(v).strip() == "":
            return 0.0
        try:
            return float(str(v).replace(",", ""))
        except ValueError:
            return 0.0

    rows: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                for row in tbl:
                    if not row or all(v is None for v in row):
                        continue

                    # Skip column-header rows
                    row_text = " ".join(str(v) for v in row if v)
                    if any(kw in row_text for kw in HEADER_KW):
                        continue
                    # Skip subtotal / grand-total rows
                    if "כ" in row_text and "הס" in row_text:
                        continue

                    # Detect passport col → employee_id is next col
                    passport_col = next(
                        (i for i, v in enumerate(row)
                         if v and PASSPORT_RE.match(str(v).strip())),
                        None,
                    )
                    if passport_col is None:
                        continue

                    emp_col = passport_col + 1
                    if emp_col >= len(row):
                        continue

                    emp_id = str(row[emp_col]).strip() if row[emp_col] else ""
                    if not EMP_ID_RE.fullmatch(emp_id):
                        continue

                    # ── Anchor-based column extraction ─────────────────────
                    # Previously we read row[0]..row[9] directly. That broke
                    # on 14/15-column variants where pdfplumber inserts a
                    # spurious None separator at a different position — every
                    # cost column would shift right by 1, with bituach (col 9
                    # nominally) actually receiving the savings_deposit value.
                    # See audit: 14/16 months had this bug, bituach was
                    # systematically 30-100% over-counted.
                    #
                    # Robust approach: walk left-to-right from col 0 to (but
                    # excluding) the passport column, collecting the first 11
                    # non-None cells in order. The 11 cost columns always sit
                    # there — only the None separators move around.
                    _cost_values = []
                    for _i in range(passport_col):
                        if len(_cost_values) >= 11: break
                        _v = row[_i]
                        if _v is None: continue
                        # Empty-string cells are sometimes used as separators
                        # too (e.g., in subtotal rows). Treat them like None.
                        if isinstance(_v, str) and _v.strip() == "": continue
                        _cost_values.append(_v)
                    # Pad with None if the row didn't yield enough values
                    # (defensive — should not happen on real data rows).
                    while len(_cost_values) < 11:
                        _cost_values.append(None)

                    rows.append({
                        "employee_id":      emp_id,
                        "employer_cost":    _n(_cost_values[0]),   # סה"כ עלות
                        "vacation_fund":    _n(_cost_values[1]),   # קרן השתלמות
                        "severance":        _n(_cost_values[2]),   # פיצויים
                        "pension":          _n(_cost_values[3]),   # פנסיה מעביד
                        "medical_insurance":_n(_cost_values[4]),   # ביטוח רפואי וחבות מעביד
                        "levy":             _n(_cost_values[5]),   # אגרות והיטלים
                        "employment_levy":  _n(_cost_values[6]),   # היטל תעסוקה ע"ז
                        "incentive_fund":   _n(_cost_values[7]),   # קרן לעידוד
                        "savings_deposit":  _n(_cost_values[8]),   # פיקדון
                        "bituach":          _n(_cost_values[9]),   # ב.לאומי מעביד
                        "gross_salary":     _n(_cost_values[10]),  # משכורת ברוטו
                    })

    if not rows:
        return EMPTY

    result = pd.DataFrame(rows)
    result["employee_id"] = result["employee_id"].astype(str).str.strip()
    result = result[result["employee_id"].str.fullmatch(r"\d{3,6}")].copy()

    _cost_cols = [
        "employer_cost", "gross_salary", "bituach", "levy", "pension",
        "vacation_fund", "severance", "medical_insurance",
        "employment_levy", "incentive_fund", "savings_deposit",
    ]
    # Sum all rows per employee (multiple site allocations)
    result = (
        result
        .groupby("employee_id", as_index=False)
        .agg({c: "sum" for c in _cost_cols})
    )
    for col in _cost_cols:
        result[col] = result[col].round(2)

    result["client"] = ""
    result["site"]   = ""

    return result.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2.  Costs from Excel
# ---------------------------------------------------------------------------

def _find_emp_id_col(df: pd.DataFrame) -> str:
    """
    Find the employee_id column by:
      1. Name matching (Hebrew / English variants)
      2. Value scan: column whose non-null values are all 3-6 digit integers
    """
    name_hints = ["מס עובד", "מספר עובד", "employee_id", "emp_id", "empid"]
    for col in df.columns:
        col_lower = str(col).lower().replace(" ", "")
        if any(h.replace(" ", "") in col_lower for h in name_hints):
            return col

    # Value-based detection: 4–6 digit integers
    for col in df.columns:
        sample = df[col].dropna().astype(str).str.strip().head(10)
        if len(sample) >= 3 and all(re.fullmatch(r"\d{3,6}", v) for v in sample):
            return col

    # Positional fallback: column index 2 (always holds emp_id in our files)
    return df.columns[2]


def _find_cost_col(df: pd.DataFrame) -> str:
    """
    Find the employer_cost column.
    Prefers last column named 'עלות' / 'cost'; falls back to last numeric column.
    """
    cost_hints = ["עלות", "cost", "employer_cost", "total_cost"]
    for col in reversed(list(df.columns)):
        col_lower = str(col).lower()
        if any(h in col_lower for h in cost_hints):
            return col
    # Fallback: last column
    return df.columns[-1]


def load_medical_deductions(hours_xls_path: str) -> pd.DataFrame:
    """
    Extract the per-employee 'ניכויי רשות - ביטוח רפואי' deduction from a
    monthly hours.xls (the old detailed Andromeda export).

    Why this is needed
    ------------------
    costs.pdf reports GROSS employer medical insurance ("ביטוח רפואי וחבות
    מעביד"). The accounting books then NET this against medical-insurance
    deductions billed back to the worker (ניכויי רשות). The net P&L line
    item — and what the user wants to see in the dashboard — equals:

        medical_insurance_net = medical_insurance(cost) - medical_deduction(hours)

    Returns
    -------
    DataFrame[employee_id (str), medical_deduction (float)]
    Empty DataFrame if the file or the column is missing — caller should
    tolerate this gracefully (treat deduction as 0 for that month).
    """
    EMPTY = pd.DataFrame(columns=["employee_id", "medical_deduction"])
    if not hours_xls_path or not str(hours_xls_path).endswith(".xls") \
            and not str(hours_xls_path).endswith(".xlsx"):
        return EMPTY
    try:
        df = pd.read_excel(hours_xls_path, sheet_name=0, header=0)
    except Exception:
        return EMPTY

    # The column name is exact in the Andromeda export — match by name to
    # survive any future column reordering.
    col_ded = "ניכויי רשות - ביטוח רפואי"
    col_emp = "מספר עובד"
    if col_ded not in df.columns or col_emp not in df.columns:
        return EMPTY

    out = pd.DataFrame({
        "employee_id":       df[col_emp].astype(str).str.strip(),
        "medical_deduction": pd.to_numeric(df[col_ded], errors="coerce").fillna(0.0),
    })
    # Keep only valid employee_ids (3-6 digit) and sum per employee in case
    # the same employee appears in multiple rows.
    out = out[out["employee_id"].str.fullmatch(r"\d{3,6}")].copy()
    out = out.groupby("employee_id", as_index=False)["medical_deduction"].sum()
    out["medical_deduction"] = out["medical_deduction"].round(2)
    return out


def load_costs_xlsx(path: str) -> pd.DataFrame:
    """
    Load employer costs from costs.xlsx.

    Known column layout (robust against garbled Hebrew headers):
      [0]  CustomerName        → client
      [1]  LocalityName        → site
      [2]  מס עובד             → employee_id
      [4]  ברוטו               → gross_salary
      [5]  ביטוח לאומי מעביד  → bituach  (employer national insurance)
      [9]  MonthlyComponent    → levy     (היטל — prorated by this module)
      [10] פנסיה מעביד         → pension
      [13] עלות                → employer_cost (total, detected dynamically)

    Returns one row per employee_id with all cost components summed.
    Multiple rows for the same employee (different site allocations) are
    collapsed via .groupby().sum() so the total is correct.
    """
    df = pd.read_excel(path, dtype=str)

    emp_id_col = _find_emp_id_col(df)
    cost_col   = _find_cost_col(df)
    client_col = df.columns[0]
    site_col   = df.columns[1]

    def _num(col_idx: int) -> pd.Series:
        if len(df.columns) > col_idx:
            return pd.to_numeric(df[df.columns[col_idx]], errors="coerce").fillna(0.0)
        return pd.Series(0.0, index=df.index)

    result = pd.DataFrame({
        "employee_id":   df[emp_id_col].astype(str).str.strip(),
        "client":        df[client_col].astype(str).str.strip(),
        "site":          df[site_col].astype(str).str.strip(),
        "employer_cost": pd.to_numeric(df[cost_col], errors="coerce").fillna(0.0),
        "gross_salary":  _num(4),
        "bituach":       _num(5),   # ביטוח לאומי מעביד
        "levy":          _num(9),   # MonthlyComponent / היטל
        "pension":       _num(10),  # פנסיה מעביד
    })

    # Keep only rows with a valid numeric employee_id
    result = result[result["employee_id"].str.fullmatch(r"\d{3,6}")].copy()

    # Sum all rows per employee — multiple rows = different site allocations
    result = (
        result
        .groupby("employee_id", as_index=False)
        .agg(
            employer_cost=("employer_cost", "sum"),
            gross_salary =("gross_salary",  "sum"),
            bituach      =("bituach",       "sum"),
            levy         =("levy",          "sum"),   # total monthly levy for employee
            pension      =("pension",       "sum"),
            client       =("client",        "first"),
            site         =("site",          "first"),
        )
    )
    for col in ("employer_cost", "gross_salary", "bituach", "levy", "pension"):
        result[col] = result[col].round(2)

    return result.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2b. Costs from a simple PDF table
# ---------------------------------------------------------------------------

def load_costs_simple_pdf(pdf_path: str) -> pd.DataFrame:
    """
    Parse a simple tabular costs PDF.

    Expected table content (column names detected by keyword, order-independent):
      - Employee ID  : מס עובד / מספר עובד / employee / id
      - Client       : לקוח / customer / client  (optional)
      - Site         : אתר / פרויקט / locality / site  (optional)
      - Total cost   : עלות / cost / total  (last matching column wins)

    One employee may span multiple rows; costs are summed per employee_id.

    Returns same schema as load_costs_xlsx:
      employee_id, employer_cost, gross_salary, bituach, levy, pension, client, site
    """
    EMPTY = pd.DataFrame(columns=[
        "employee_id", "employer_cost", "gross_salary",
        "bituach", "levy", "pension", "client", "site",
    ])

    try:
        import pdfplumber
    except ImportError:
        return EMPTY

    EMP_ID_RE = re.compile(r"^\d{3,6}$")

    def _to_float(v: object) -> float:
        if v is None or str(v).strip() in ("", "-"):
            return 0.0
        try:
            return float(str(v).replace(",", "").replace("₪", "").strip())
        except ValueError:
            return 0.0

    def _col_idx(header_row: list, hints: list[str]) -> int | None:
        """Return the last column index whose header matches any hint."""
        found = None
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            cell_n = str(cell).lower().replace(" ", "").replace('"', "")
            if any(h in cell_n for h in hints):
                found = i
        return found

    ID_HINTS     = ["מסעובד", "מספרעובד", "employee", "empid"]
    COST_HINTS   = ["עלות", "cost", "total"]
    CLIENT_HINTS = ["לקוח", "customer", "client"]
    SITE_HINTS   = ["אתר", "פרויקט", "locality", "site"]
    SKIP_KW      = {"סהכ", "סה\"כ", "total", "ס\"כ"}

    rows: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                if not tbl or len(tbl) < 2:
                    continue

                header = tbl[0]

                id_col     = _col_idx(header, ID_HINTS)
                cost_col   = _col_idx(header, COST_HINTS)
                client_col = _col_idx(header, CLIENT_HINTS)
                site_col   = _col_idx(header, SITE_HINTS)

                for row in tbl[1:]:
                    if not row:
                        continue

                    if id_col is not None and cost_col is not None:
                        # Header-guided extraction
                        raw_id = str(row[id_col] or "").strip() if id_col < len(row) else ""
                        if not EMP_ID_RE.fullmatch(raw_id):
                            continue
                        # Skip subtotal rows
                        row_text = " ".join(str(v) for v in row if v)
                        if any(kw in row_text for kw in SKIP_KW):
                            continue
                        rows.append({
                            "employee_id":   raw_id,
                            "employer_cost": _to_float(row[cost_col]) if cost_col < len(row) else 0.0,
                            "client": str(row[client_col] or "").strip() if client_col is not None and client_col < len(row) else "",
                            "site":   str(row[site_col]   or "").strip() if site_col   is not None and site_col   < len(row) else "",
                        })
                    else:
                        # Positional fallback: scan each cell for a 3-6 digit ID
                        row_text = " ".join(str(v) for v in row if v)
                        if any(kw in row_text for kw in SKIP_KW):
                            continue
                        for ci, cell in enumerate(row):
                            if cell and EMP_ID_RE.fullmatch(str(cell).strip()):
                                emp_id = str(cell).strip()
                                # Find largest numeric value in remaining cells → cost
                                best_cost, best_j = 0.0, -1
                                for cj, val in enumerate(row):
                                    if cj == ci:
                                        continue
                                    try:
                                        v = float(str(val or "").replace(",", "").replace("₪", ""))
                                        if v > best_cost:
                                            best_cost, best_j = v, cj
                                    except (ValueError, TypeError):
                                        pass
                                if best_cost > 0:
                                    rows.append({
                                        "employee_id":   emp_id,
                                        "employer_cost": best_cost,
                                        "client": str(row[0] or "").strip() if row else "",
                                        "site":   str(row[1] or "").strip() if len(row) > 1 else "",
                                    })
                                break

    if not rows:
        return EMPTY

    result = pd.DataFrame(rows)
    result["employee_id"] = result["employee_id"].astype(str).str.strip()
    result = result[result["employee_id"].str.fullmatch(r"\d{3,6}")].copy()

    result = (
        result.groupby("employee_id", as_index=False)
        .agg(
            employer_cost=("employer_cost", "sum"),
            client       =("client",        "first"),
            site         =("site",          "first"),
        )
    )
    result["employer_cost"] = result["employer_cost"].round(2)
    result["gross_salary"]  = 0.0
    result["bituach"]       = 0.0
    result["levy"]          = 0.0
    result["pension"]       = 0.0

    return result.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 3.  Merge + cost allocation
# ---------------------------------------------------------------------------

def merge_and_allocate(
    hours_df: pd.DataFrame,
    costs_df: pd.DataFrame,
    month: str,
    worked_days_series: "pd.Series | None" = None,
    std_days_per_month: int = 22,
    std_hours_per_day: float = 8.0,
) -> pd.DataFrame:
    """
    Join hours and costs by employee_id and allocate cost across sites.

    Algorithm
    ---------
    1. Aggregate total hours per employee (across ALL sites this month).
    2. Recalculate levy proportionally by days worked (see below).
    3. Rebuild employer_cost = original − raw_levy + adjusted_levy.
    4. cost_per_hour = employer_cost / emp_total_hours  (employee level).
    5. allocated_cost = site_hours × cost_per_hour  (per site row).

    Levy (אגרות) recalculation
    --------------------------
    The raw levy from the source is the FULL monthly fee.  Employees who
    worked fewer than std_days_per_month days receive a proportional share:

        worked_days  = work_days column (from hours file)
                       OR total_hours / std_hours_per_day when unavailable
        adjusted_levy = (worked_days / std_days_per_month) × raw_levy
        adjusted_levy = min(adjusted_levy, raw_levy)   ← capped at full fee

    Parameters
    ----------
    std_days_per_month  Full working-month reference (default 22).
    std_hours_per_day   Hours per standard working day (default 8); used only
                        when the hours file has no daily date breakdown.
    """
    if hours_df.empty:
        return pd.DataFrame()

    h = hours_df.copy()
    c = costs_df.copy()

    h["employee_id"] = h["employee_id"].astype(str).str.strip()
    c["employee_id"] = c["employee_id"].astype(str).str.strip()

    # Step 1: total hours AND days per employee (across all sites)
    emp_totals = (
        h.groupby("employee_id", as_index=False)
        .agg(emp_total_hours=("total_hours", "sum"),
             emp_total_days =("work_days",   "sum"))
    )

    # Left join: all hour rows kept; missing cost rows → cost columns = 0
    # If hours already carries 'client' (e.g. from billing xlsx), don't
    # pull client from costs to avoid a _x/_y column collision.
    _all_cost_cols = [
        "employer_cost", "gross_salary", "bituach", "levy", "pension",
        "vacation_fund", "severance", "medical_insurance",
        "employment_levy", "incentive_fund", "savings_deposit",
    ]
    cost_cols = [c2 for c2 in ["employee_id", "client"] + _all_cost_cols
                 if c2 in c.columns and not (c2 == "client" and "client" in h.columns)]
    merged = (
        h
        .merge(emp_totals, on="employee_id", how="left")
        .merge(c[cost_cols],  on="employee_id", how="left")
    )
    for col in _all_cost_cols:
        if col in merged.columns:
            merged[col] = merged[col].fillna(0.0)

    # ── Levy (אגרות) recalculation — day-based proration ─────────────────────
    #
    # Rule:
    #   • The levy values in costs.pdf are unreliable for partial-month workers
    #     (accounting system sometimes records a prorated or incorrect amount).
    #   • Standard monthly levy = median levy of full-month workers (22+ days)
    #     with levy > 500 (ignores anomalous low values). This is the official
    #     government fee for a full working month.
    #   • adjusted_levy = standard_levy × (worked_days / 22), capped at standard_levy.

    if "levy" in merged.columns:
        emp_work_days = merged.groupby("employee_id")["emp_total_days"].first()

        if worked_days_series is not None:
            actual_days = (
                worked_days_series
                .reindex(emp_work_days.index)
                .fillna(emp_work_days)
            )
        else:
            emp_total_h = merged.groupby("employee_id")["emp_total_hours"].first()
            actual_days = emp_work_days.where(
                emp_work_days > 0,
                emp_total_h / std_hours_per_day,
            )

        # Determine standard full-month levy from employees who worked 22+ days
        emp_raw_levy  = merged.groupby("employee_id")["levy"].first()
        full_month_ids = emp_work_days[emp_work_days >= std_days_per_month].index
        full_month_levies = emp_raw_levy.loc[full_month_ids]
        full_month_levies = full_month_levies[full_month_levies > 500]  # drop anomalous low values

        if not full_month_levies.empty:
            standard_levy = float(full_month_levies.median())
        else:
            # Fallback: use median of all levy > 500 in this month
            all_levies = emp_raw_levy[emp_raw_levy > 500]
            standard_levy = float(all_levies.median()) if not all_levies.empty else 0.0

        # Apply standard levy prorated by actual days (22 = full month)
        adjusted_levy_map = (
            (actual_days / std_days_per_month).clip(upper=1.0) * standard_levy
        ).round(2)

        merged = merged.merge(
            adjusted_levy_map.rename("adjusted_levy").reset_index(),
            on="employee_id", how="left",
        )
        merged["adjusted_levy"] = merged["adjusted_levy"].fillna(0.0)

        # Rebuild employer_cost: swap raw levy for adjusted levy
        merged["employer_cost"] = (
            merged["employer_cost"] - merged["levy"] + merged["adjusted_levy"]
        ).clip(lower=0).round(2)

        # Expose the raw full-month levy for reference / validation
        merged = merged.merge(
            emp_raw_levy.rename("full_monthly_levy").reset_index(),
            on="employee_id", how="left",
        )
    else:
        merged["adjusted_levy"]    = 0.0
        merged["full_monthly_levy"] = 0.0

    # ── Cost allocation across sites ──────────────────────────────────────────
    # Step 2: cost_per_hour = employer_cost / emp_total_hours  (employee-level)
    _safe_total = merged["emp_total_hours"].replace(0, float("nan"))
    merged["cost_per_hour"] = (
        merged["employer_cost"] / _safe_total
    ).round(4).fillna(0.0)

    # Step 3: allocated_cost = site_hours × cost_per_hour
    merged["allocated_cost"] = (
        merged["total_hours"] * merged["cost_per_hour"]
    ).round(2).fillna(0.0)

    # Step 4: allocate each cost component proportionally to site hours
    _alloc_ratio = (merged["total_hours"] / _safe_total).fillna(0.0)
    _extra_comps = ["vacation_fund", "severance", "medical_insurance",
                    "employment_levy", "incentive_fund", "savings_deposit"]
    for _comp in _extra_comps:
        if _comp in merged.columns:
            merged[f"{_comp}_alloc"] = (merged[_comp] * _alloc_ratio).round(2)

    merged["cost_per_hour"] = merged["cost_per_hour"].round(2)
    merged["month"] = month

    return merged.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 4.  Build output sheets
# ---------------------------------------------------------------------------

def build_sheets(merged: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Build the four DataFrames for cost_analysis.xlsx.

    Keys: 'employee_cost', 'site_cost', 'client_cost'
    """
    if merged.empty:
        empty = pd.DataFrame()
        return {
            "employee_summary": empty,
            "site_summary":     empty,
            "client_summary":   empty,
        }

    # ── Sheet 1: employee_summary ────────────────────────────────────────────
    # employer_cost is already summed per employee_id in load_costs_xlsx()
    agg_spec: dict = {
        "employee_name": ("employee_name",  "first"),
        "total_hours"  : ("total_hours",    "sum"),
        "employer_cost": ("employer_cost",  "first"),  # identical across sites
    }
    if "country" in merged.columns:
        agg_spec["country"] = ("country", "first")
    if "adjusted_levy" in merged.columns:
        agg_spec["adjusted_levy"] = ("adjusted_levy", "first")

    emp_cost = merged.groupby(["month", "employee_id"], as_index=False).agg(**agg_spec)

    _safe = emp_cost["total_hours"].replace(0, float("nan"))
    emp_cost["cost_per_hour"] = (
        emp_cost["employer_cost"] / _safe
    ).round(2).fillna(0.0)
    emp_cost["total_hours"]   = emp_cost["total_hours"].round(2)
    emp_cost["employer_cost"] = emp_cost["employer_cost"].round(2)

    out_cols = ["month", "employee_id", "employee_name"]
    if "country" in emp_cost.columns:
        out_cols.append("country")
    out_cols += ["total_hours", "employer_cost", "cost_per_hour"]
    if "adjusted_levy" in emp_cost.columns:
        emp_cost["adjusted_levy"] = emp_cost["adjusted_levy"].round(2)
        out_cols.append("adjusted_levy")

    emp_cost = emp_cost[out_cols].sort_values(
        ["month", "employee_id"]
    ).reset_index(drop=True)

    # ── Sheet 2: site_summary ─────────────────────────────────────────────────
    site_cost_cols = [c for c in
        ["month", "site", "employee_id", "country", "total_hours",
         "allocated_cost", "cost_per_hour"]
        if c in merged.columns
    ]
    site_cost = merged[site_cost_cols].copy().rename(columns={"total_hours": "hours"})
    site_cost = site_cost.sort_values(
        ["month", "site", "employee_id"]
    ).reset_index(drop=True)

    # ── Sheet 3: client_summary ───────────────────────────────────────────────
    if "client" in merged.columns:
        client_cost = (
            merged
            .dropna(subset=["client"])
            .query("client != ''")
            .groupby(["month", "client"], as_index=False)
            .agg(
                total_hours=("total_hours",    "sum"),
                total_cost =("allocated_cost", "sum"),
            )
        )
        _safe_c = client_cost["total_hours"].replace(0, float("nan"))
        client_cost["avg_cost_per_hour"] = (
            client_cost["total_cost"] / _safe_c
        ).round(2).fillna(0.0)
        client_cost["total_hours"] = client_cost["total_hours"].round(2)
        client_cost["total_cost"]  = client_cost["total_cost"].round(2)
        client_cost = client_cost[
            ["month", "client", "total_hours", "total_cost", "avg_cost_per_hour"]
        ].sort_values(["month", "total_cost"], ascending=[True, False]).reset_index(drop=True)
    else:
        client_cost = pd.DataFrame(
            columns=["month", "client", "total_hours", "total_cost", "avg_cost_per_hour"]
        )

    return {
        "employee_summary": emp_cost,
        "site_summary":     site_cost,
        "client_summary":   client_cost,
    }


# ---------------------------------------------------------------------------
# 5.  Warnings / validation
# ---------------------------------------------------------------------------

def detect_warnings(
    hours_df: pd.DataFrame,
    costs_df: pd.DataFrame,
    merged: pd.DataFrame,
    month: str,
    cost_per_hour_threshold: float = 250.0,
) -> pd.DataFrame:
    """
    Return a DataFrame of data quality issues.

    Checks:
      1. Employee in hours PDF but missing from costs.xlsx → cost set to 0
      2. Employee has zero hours this month with positive cost → excluded, warned
      3. cost_per_hour > threshold (default 250)
    """
    rows: list[dict] = []

    def _warn(emp_id: str, issue: str) -> None:
        rows.append({"month": month, "employee_id": emp_id, "issue": issue})

    if hours_df.empty:
        return pd.DataFrame(columns=["month", "employee_id", "issue"])

    hours_ids = set(hours_df["employee_id"].astype(str).str.strip())
    costs_ids = set(costs_df["employee_id"].astype(str).str.strip()) if not costs_df.empty else set()

    # 1. In hours but not in costs → kept with employer_cost = 0
    for eid in sorted(hours_ids - costs_ids):
        _warn(eid, "Missing cost data — employer_cost set to 0")

    if not merged.empty:
        # 2. Zero hours with positive employer cost (edge case)
        zero_h = merged[(merged["total_hours"] == 0) & (merged["employer_cost"] > 0)]
        for _, r in zero_h.drop_duplicates("employee_id").iterrows():
            _warn(
                str(r["employee_id"]),
                f"Zero hours but employer_cost ₪{r['employer_cost']:,.0f} — cost_per_hour set to 0",
            )

        # 3. Unusually high cost_per_hour
        high_rate = merged[merged["cost_per_hour"] > cost_per_hour_threshold]
        for _, r in high_rate.drop_duplicates("employee_id").iterrows():
            _warn(
                str(r["employee_id"]),
                f"High cost_per_hour ₪{r['cost_per_hour']:.0f}/h "
                f"(threshold ₪{cost_per_hour_threshold:.0f}/h)",
            )

        # 4. No negative values in allocated_cost or cost_per_hour
        neg = merged[
            (merged["allocated_cost"] < 0) | (merged["cost_per_hour"] < 0)
        ]
        for _, r in neg.drop_duplicates("employee_id").iterrows():
            _warn(
                str(r["employee_id"]),
                f"Negative value: allocated_cost={r['allocated_cost']:.2f}, "
                f"cost_per_hour={r['cost_per_hour']:.2f}",
            )

        # 5. Sum of site hours per employee must equal emp_total_hours
        hours_check = (
            merged.groupby("employee_id", as_index=True)
            .agg(sum_site=("total_hours",    "sum"),
                 emp_total=("emp_total_hours","first"))
        )
        bad_hours = hours_check[
            (hours_check["sum_site"] - hours_check["emp_total"]).abs() > 0.01
        ]
        for eid, r in bad_hours.iterrows():
            _warn(
                str(eid),
                f"Hours mismatch: sum(site_hours)={r['sum_site']:.2f} "
                f"≠ emp_total_hours={r['emp_total']:.2f}",
            )

        # 6. Levy validation — adjusted_levy must not exceed full_monthly_levy
        if "adjusted_levy" in merged.columns and "full_monthly_levy" in merged.columns:
            over = merged[
                merged["adjusted_levy"] > merged["full_monthly_levy"] + 0.01
            ]
            for _, r in over.drop_duplicates("employee_id").iterrows():
                _warn(str(r["employee_id"]),
                      f"adjusted_levy ₪{r['adjusted_levy']:.2f} > "
                      f"full_monthly_levy ₪{r['full_monthly_levy']:.2f}")

            neg_lv = merged[merged["adjusted_levy"] < 0]
            for _, r in neg_lv.drop_duplicates("employee_id").iterrows():
                _warn(str(r["employee_id"]),
                      f"Negative adjusted_levy ₪{r['adjusted_levy']:.2f}")

        # 7. Allocation validation: sum(allocated_cost) per employee must equal employer_cost
        alloc_check = (
            merged.groupby("employee_id", as_index=False)
            .agg(sum_allocated=("allocated_cost", "sum"),
                 employer_cost =("employer_cost",  "first"))
        )
        alloc_check["diff"] = (
            alloc_check["sum_allocated"] - alloc_check["employer_cost"]
        ).abs()
        mismatch = alloc_check[alloc_check["diff"] > 0.05]
        for _, r in mismatch.iterrows():
            _warn(
                str(r["employee_id"]),
                f"Allocation mismatch: sum(allocated) ₪{r['sum_allocated']:,.2f} "
                f"≠ employer_cost ₪{r['employer_cost']:,.2f} "
                f"(diff ₪{r['diff']:,.2f})",
            )

    df = pd.DataFrame(rows, columns=["month", "employee_id", "issue"])
    return df.drop_duplicates().reset_index(drop=True)


# ---------------------------------------------------------------------------
# 6.  Excel export — plain data
# ---------------------------------------------------------------------------

def export_to_excel(
    output_path: str,
    sheets: dict[str, pd.DataFrame],
    warnings_df: pd.DataFrame,
) -> None:
    """Write cost_analysis.xlsx with 4 data sheets (plain format)."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name in ["employee_summary", "site_summary", "client_summary"]:
            sheets.get(name, pd.DataFrame()).to_excel(
                writer, sheet_name=name, index=False
            )
        warnings_df.to_excel(writer, sheet_name="warnings", index=False)


# ---------------------------------------------------------------------------
# 6b. Hebrew professional Excel report
# ---------------------------------------------------------------------------

def export_hebrew_report(
    output_path: str,
    sheets: dict[str, pd.DataFrame],
    warnings_df: pd.DataFrame,
    title_suffix: str = "",
) -> None:
    """
    Write a professional, Hebrew-formatted Excel cost report.

    Sheets (Hebrew names):
      סיכום עובדים | סיכום אתרים | סיכום לקוחות | בדיקות ובעיות

    Features:
      - Hebrew column headers
      - Styled header row (bold, centered, blue background)
      - Currency format for money columns
      - Number format for hours
      - Auto column width
      - Freeze top rows (title + header)
      - Borders on all data cells
      - TOTAL row at bottom of each data sheet
      - Right-aligned numbers, center-aligned text
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    # ── Style constants ───────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue
    TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
    STRIPE_FILL  = PatternFill("solid", fgColor="EBF3FB")   # very light blue
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT   = Font(bold=True, color="FFFFFF", size=13)
    TOTAL_FONT   = Font(bold=True, size=11)
    DATA_FONT    = Font(size=10)
    THIN         = Side(style="thin", color="B0C4DE")
    THICK        = Side(style="medium", color="1F4E79")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    OUTER_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

    ILS_FMT   = '₪#,##0.00'
    ILS_FMT0  = '₪#,##0'
    HRS_FMT   = '#,##0.00'
    INT_FMT   = '#,##0'

    RIGHT = Alignment(horizontal="right",  vertical="center")
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT_WRAP = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def _border_cell(ws, row, col, value, fmt=None, font=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        c.border = BORDER
        if fmt:    c.number_format = fmt
        if font:   c.font = font
        if fill:   c.fill = fill
        if align:  c.alignment = align
        return c

    def _auto_width(ws, min_w=8, max_w=45):
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0)
                for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                min(max_w, max(min_w, max_len + 2))
            )

    def _title_row(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=text)
        c.font  = TITLE_FONT
        c.fill  = TITLE_FILL
        c.alignment = CENTER
        ws.row_dimensions[1].height = 24

    def _header_row(ws, headers, row=2):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font      = HEADER_FONT
            c.fill      = HEADER_FILL
            c.border    = BORDER
            c.alignment = RIGHT_WRAP
        ws.row_dimensions[row].height = 30

    def _write_data(ws, df, start_row, col_defs):
        """
        col_defs: list of (series_col, fmt, align, do_total) tuples.
        do_total=False → column omitted from the TOTAL row.
        Returns (last_data_row, totals_dict).
        """
        totals: dict[int, float] = {}
        for ri, (_, row_data) in enumerate(df.iterrows()):
            r = start_row + ri
            fill = STRIPE_FILL if ri % 2 == 0 else None
            for ci, tup in enumerate(col_defs, 1):
                series_col, fmt, align = tup[0], tup[1], tup[2]
                do_total = tup[3] if len(tup) > 3 else True
                val = row_data[series_col] if series_col is not None else None
                _border_cell(ws, r, ci, val, fmt=fmt, font=DATA_FONT,
                             fill=fill, align=align or RIGHT)
                if do_total and isinstance(val, (int, float)) and fmt:
                    totals[ci] = totals.get(ci, 0.0) + (val or 0)
        return start_row + len(df) - 1, totals

    def _total_row(ws, row, col_defs, totals):
        ws.cell(row=row, column=1, value="סה\"כ").font = TOTAL_FONT
        ws.cell(row=row, column=1).fill = TOTAL_FILL
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=1).alignment = CENTER
        for ci in range(2, len(col_defs) + 1):
            tup = col_defs[ci - 1]
            fmt = tup[1] if len(tup) > 1 else None
            val = totals.get(ci)   # None if not summed
            _border_cell(ws, row, ci, val, fmt=fmt,
                         font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT)
        ws.row_dimensions[row].height = 18

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    title_base = f"דוח עלויות עובדים"
    if title_suffix:
        title_base += f" — {title_suffix}"

    # ── Sheet 1: סיכום עובדים ─────────────────────────────────────────────────
    emp = sheets.get("employee_summary", pd.DataFrame()).copy()
    if not emp.empty:
        emp = emp.sort_values("cost_per_hour", ascending=False)
        emp = emp[emp["employer_cost"] > 0]

    ws1 = wb.create_sheet("סיכום עובדים")
    ws1.sheet_view.rightToLeft = True
    hdr1 = ["חודש", "מספר עובד", "שם עובד", 'סה"כ שעות', "עלות מעביד", "אגרות", "עלות לשעה"]
    _title_row(ws1, title_base, len(hdr1))
    _header_row(ws1, hdr1, row=2)

    col_defs1 = [
        ("month",           None,     CENTER,  False),
        ("employee_id",     None,     CENTER,  False),
        ("employee_name",   None,     CENTER,  False),
        ("total_hours",     HRS_FMT,  RIGHT,   True),
        ("employer_cost",   ILS_FMT0, RIGHT,   True),
        ("adjusted_levy", ILS_FMT0, RIGHT,   True) if "adjusted_levy" in emp.columns
            else (None,             None,     CENTER, False),
        ("cost_per_hour",   ILS_FMT0, RIGHT,   False),
    ]

    last_r1, totals1 = _write_data(ws1, emp, start_row=3, col_defs=col_defs1)
    _total_row(ws1, last_r1 + 1, col_defs1, totals1)
    ws1.freeze_panes = "A3"
    _auto_width(ws1)

    # ── Sheet 2: סיכום אתרים ──────────────────────────────────────────────────
    site = sheets.get("site_summary", pd.DataFrame()).copy()
    if not site.empty:
        site = site.sort_values("allocated_cost", ascending=False)
        site = site[site["allocated_cost"] > 0]

    ws2 = wb.create_sheet("סיכום אתרים")
    ws2.sheet_view.rightToLeft = True
    hdr2 = ["חודש", "אתר", "מספר עובד", "שעות", "עלות", "עלות לשעה"]
    _title_row(ws2, title_base, len(hdr2))
    _header_row(ws2, hdr2, row=2)
    col_defs2 = [
        ("month",          None,     CENTER,  False),
        ("site",           None,     CENTER,  False),
        ("employee_id",    None,     CENTER,  False),
        ("hours",          HRS_FMT,  RIGHT,   True),
        ("allocated_cost", ILS_FMT0, RIGHT,   True),
        ("cost_per_hour",  ILS_FMT0, RIGHT,   False),
    ]
    last_r2, totals2 = _write_data(ws2, site, start_row=3, col_defs=col_defs2)
    _total_row(ws2, last_r2 + 1, col_defs2, totals2)
    ws2.freeze_panes = "A3"
    _auto_width(ws2)

    # ── Sheet 3: סיכום לקוחות ────────────────────────────────────────────────
    cli = sheets.get("client_summary", pd.DataFrame()).copy()
    if not cli.empty:
        cli = cli[cli["client"].astype(str).str.strip() != ""]
        cli = cli.sort_values("total_cost", ascending=False)

    ws3 = wb.create_sheet("סיכום לקוחות")
    ws3.sheet_view.rightToLeft = True
    hdr3 = ["חודש", "לקוח", 'סה"כ שעות', 'סה"כ עלות', "עלות ממוצעת לשעה"]
    _title_row(ws3, title_base, len(hdr3))
    _header_row(ws3, hdr3, row=2)
    col_defs3 = [
        ("month",             None,     CENTER,  False),
        ("client",            None,     CENTER,  False),
        ("total_hours",       HRS_FMT,  RIGHT,   True),
        ("total_cost",        ILS_FMT0, RIGHT,   True),
        ("avg_cost_per_hour", ILS_FMT0, RIGHT,   False),
    ]
    last_r3, totals3 = _write_data(ws3, cli, start_row=3, col_defs=col_defs3)
    _total_row(ws3, last_r3 + 1, col_defs3, totals3)
    ws3.freeze_panes = "A3"
    _auto_width(ws3)

    # ── Sheet 4: בדיקות ובעיות ───────────────────────────────────────────────
    ws4 = wb.create_sheet("בדיקות ובעיות")
    ws4.sheet_view.rightToLeft = True
    hdr4 = ["חודש", "מספר עובד", "תיאור בעיה"]
    _title_row(ws4, title_base, len(hdr4))
    _header_row(ws4, hdr4, row=2)
    warn = warnings_df.copy()
    for ri, (_, row_data) in enumerate(warn.iterrows()):
        r = 3 + ri
        fill = STRIPE_FILL if ri % 2 == 0 else None
        _border_cell(ws4, r, 1, row_data.get("month"),       font=DATA_FONT, fill=fill, align=CENTER)
        _border_cell(ws4, r, 2, row_data.get("employee_id"), font=DATA_FONT, fill=fill, align=CENTER)
        _border_cell(ws4, r, 3, row_data.get("issue"),       font=DATA_FONT, fill=fill, align=RIGHT)
    if warn.empty:
        ws4.cell(row=3, column=1, value="✓ אין בעיות").font = Font(color="00AA00", bold=True)
    ws4.freeze_panes = "A3"
    _auto_width(ws4)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# 7.  High-level runner (single month)
# ---------------------------------------------------------------------------

def run_month(
    pdf_path: str,
    costs_path: str,
    month: str,
    output_path: str,
    cost_per_hour_threshold: float = 250.0,
) -> dict:
    """
    Full pipeline for one month:
      parse hours (PDF or xlsx) → load costs → merge → build sheets → export

    Returns a summary dict for the CLI to print.
    """
    if pdf_path.lower().endswith(".xlsx"):
        hours_df = load_hours_from_xlsx(pdf_path, month)
    else:
        hours_df = load_hours_from_pdf(pdf_path, month)
    worked_days_s = compute_worked_days(pdf_path)
    costs_df = load_costs_xlsx(costs_path)
    merged   = merge_and_allocate(hours_df, costs_df, month,
                                  worked_days_series=worked_days_s)
    sheets   = build_sheets(merged)
    warnings = detect_warnings(hours_df, costs_df, merged, month, cost_per_hour_threshold)
    export_to_excel(output_path, sheets, warnings)

    # Summary stats
    emp_cost = sheets.get("employee_summary", pd.DataFrame())
    return {
        "month":           month,
        "employees_hours": hours_df["employee_id"].nunique() if not hours_df.empty else 0,
        "employees_costs": len(costs_df),
        "employees_merged": merged["employee_id"].nunique() if not merged.empty else 0,
        "total_hours":     round(float(hours_df["total_hours"].sum()), 2) if not hours_df.empty else 0,
        "total_cost":      round(float(costs_df["employer_cost"].sum()), 2) if not costs_df.empty else 0,
        "avg_cost_per_hour": round(
            float(emp_cost["cost_per_hour"].mean()), 2
        ) if not emp_cost.empty else 0,
        "warnings":        len(warnings),
        "output_path":     output_path,
    }
