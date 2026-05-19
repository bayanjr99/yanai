"""
core/standards_loader.py — טעינה ועיבוד קובץ תקן.xlsx (59 שורות).

עמודות הקובץ:
  A=מספר בחשבשבת, B=לקוח, C=אתר, D=מדינה,
  E=סוג ("שעות"/"ימים"), F=מחיר, G=הפסקות, H=תקן

billing_kind values:
  hourly_no_completion   — שעתי, ללא השלמה חודשית
  hourly_with_completion — שעתי + השלמה למינימום (236h / 220h)
  daily_with_ot          — יומי + שעות נוספות מעל X/יום
  daily_no_ot            — יומי בלי שעות נוספות
  daily_or_monthly_min   — יומי 10ש' ביום חול / 236 לפי הנמוך (ולפמן)
  daily_min_only         — מינימום שעות ביום (שמואל שמעון 10ש'/יום)
  mixed                  — גם יומי וגם שעתי לאותו אתר (בראל, ח.י יוסף, צלעג)
  unknown                — לא ניתן לסווג
  missing_data           — חסרים נתונים (מחיר/סוג)
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

_BILLING_TYPE_MAP = {
    "hourly":            "hourly_no_completion",
    "hourly_completion": "hourly_with_completion",
    "daily":             "daily_with_ot",
    "unknown":           "unknown",
}

_KIND_DISPLAY = {
    "hourly_no_completion":   "⏱️ שעתי",
    "hourly_with_completion": "⏱️ שעתי + השלמה",
    "daily_with_ot":          "📅 יומי + OT",
    "daily_no_ot":            "📅 יומי בלבד",
    "daily_or_monthly_min":   "🔀 יומי / חודשי",
    "daily_min_only":         "📅 מינ' יומי",
    "mixed":                  "🔀 מעורב",
    "unknown":                "❓ לא מוגדר",
    "missing_data":           "⚠️ חסרים נתונים",
}


def _parse_completion(
    text: str,
    billing_type: str,
    rate: float | None,
) -> tuple[str, int | None, int | None, float | None]:
    """
    מחזיר (billing_kind, completion_target_hours, daily_min_hours, ot_hourly_rate).
    """
    t = str(text or "").strip()
    none_like = not t or t in ("-", "nan", "None")

    if none_like:
        if billing_type == "daily":
            return ("daily_no_ot", None, None, None)
        return ("hourly_no_completion", None, None, None)

    # 'מחיר יומי עד X שעות, כל שעה נוספת לפי Y ₪'
    m = re.search(r"יומי\s*עד\s*(\d+)\s*שעות.*?(\d+(?:\.\d+)?)\s*[₪]", t)
    if m:
        return ("daily_with_ot", None, int(m.group(1)), float(m.group(2)))

    # 'תקן יומי: 10 שעות ביום חול, X שעות ביום שישי או 236 שעות לפי הנמוך'
    if "תקן יומי" in t and re.search(r"23[0-9]", t):
        m_day = re.search(r"(\d+)\s*שעות ביום חול", t)
        daily_min = int(m_day.group(1)) if m_day else 10
        return ("daily_or_monthly_min", 236, daily_min, rate)

    # 'השלמה ל-N שעות/ל-N': N ≤ 12 = מינימום יומי; N ≥ 100 = מינימום חודשי
    m = re.search(r"השלמה\s*ל-?\s*(\d+)", t)
    if m:
        target = int(m.group(1))
        if target <= 12:
            return ("daily_min_only", None, target, rate)
        return ("hourly_with_completion", target, None, None)

    # 'ללא שעות נוספות' / 'ללא OT'
    if "ללא שעות נוספות" in t or "ללא OT" in t.upper():
        return ("daily_no_ot", None, None, None)

    # יומי generic
    if billing_type == "daily" or "יומי" in t:
        return ("daily_with_ot", None, 10, rate)

    return ("unknown", None, None, None)


def load_standards(data_dir: Path | str = "data") -> pd.DataFrame:
    """
    קורא data/תקן.xlsx ומחזיר DataFrame (שורה לכל אתר/לקוח).

    עמודות:
      account_code, client_full, site, country, billing_type,
      rate, include_breaks, completion_raw, billing_kind,
      completion_target_hours, daily_min_hours, ot_hourly_rate, is_complete

    Loudly logs a warning when the file is missing or unreadable — silent
    fallback to an empty DataFrame previously hid this from operators and the
    dashboard would display zero expected billing for every client.
    """
    import logging
    _log = logging.getLogger("standards_loader")

    data_dir = Path(data_dir)
    if not data_dir.exists():
        _log.warning("STANDARDS FILE MISSING — data directory not found: %s", data_dir)
        return pd.DataFrame()

    std_file: Path | None = None
    try:
        for f in data_dir.iterdir():
            if f.is_file() and f.suffix.lower() == ".xlsx" and "תקן" in f.name:
                std_file = f
                break
    except Exception as exc:
        _log.warning("STANDARDS FILE MISSING — cannot enumerate %s: %s", data_dir, exc)
        return pd.DataFrame()

    if std_file is None:
        _log.warning("STANDARDS FILE MISSING — no תקן.xlsx in %s. "
                     "All billing will be marked 'unknown'.", data_dir)
        return pd.DataFrame()

    try:
        wb = openpyxl.load_workbook(str(std_file), data_only=True)
    except Exception as exc:
        _log.error("STANDARDS FILE CORRUPT — cannot open %s: %s", std_file, exc)
        print(f"[standards] Cannot open {std_file}: {exc}")
        return pd.DataFrame()

    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not any(r):
            continue

        account = str(r[0]).strip() if r[0] is not None else ""
        client  = str(r[1]).strip() if r[1] is not None else ""
        site    = str(r[2]).strip() if r[2] is not None else ""
        country = str(r[3]).strip() if r[3] is not None else ""
        kind_raw = r[4]
        price    = r[5]
        breaks   = str(r[6]).strip() if r[6] is not None else "-"
        comp_raw = str(r[7]).strip() if r[7] is not None else "-"

        if not client or client == "nan":
            continue

        # זיהוי שורות עם תזוזת עמודות (R19 — price בעמודה E, site ריק)
        if isinstance(kind_raw, (int, float)) and price is None:
            price    = kind_raw
            kind_raw = "שעות"   # default reasonable

        # is_complete: יש מחיר וסוג
        kind_str = str(kind_raw or "").strip()
        has_kind  = kind_str in ("שעות", "ימים")
        has_price = isinstance(price, (int, float))
        is_complete = has_kind and has_price

        billing_type = "daily" if "ימי" in kind_str else "hourly"

        billing_kind, comp_target, daily_min, ot_rate = _parse_completion(
            comp_raw, billing_type, float(price) if has_price else None
        )

        if not is_complete:
            billing_kind = "missing_data"

        include_breaks = (
            breaks.upper() == "V"
            or "שעה נוספת" in breaks
            or "לשעה" in breaks
        )

        rows.append({
            "account_code":            account,
            "client_full":             client,
            "site":                    site or "כל האתרים",
            "country":                 country,
            "billing_type":            billing_type,
            "rate":                    float(price) if has_price else None,
            "include_breaks":          include_breaks,
            "completion_raw":          comp_raw,
            "billing_kind":            billing_kind,
            "completion_target_hours": comp_target,
            "daily_min_hours":         daily_min,
            "ot_hourly_rate":          ot_rate,
            "is_complete":             is_complete,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # סמן mixed: אתרים שיש להם גם ימים וגם שעות
    _site_kinds = (
        df[df["is_complete"]]
        .groupby(["client_full", "site"])["billing_type"]
        .nunique()
    )
    _mixed_sites = _site_kinds[_site_kinds > 1].index
    for client_f, site_s in _mixed_sites:
        mask = (df["client_full"] == client_f) & (df["site"] == site_s)
        df.loc[mask, "billing_kind"] = "mixed"

    return df


# Country spelling variants seen in either תקן.xlsx or Andromeda data —
# normalised to the canonical form used in `processed_data`. This guards
# against typos and grammatical variants (singular vs plural, etc.) that
# would otherwise cause silent rate mismatches.
_COUNTRY_ALIASES: dict[str, str] = {
    # Moldovan variants → "מולדובה" (the canonical country name as it
    # appears in the Andromeda data)
    "מודובנים":   "מולדובה",
    "מודובנם":    "מולדובה",
    "מולדובים":   "מולדובה",
    "מולדבים":    "מולדובה",
    "מולדבנים":   "מולדובה",
    "מולדבני":    "מולדובה",
    # Thai
    "תאילנדים":   "תאילנד",
    # Indian
    "הודים":      "הודו",
    # Sri Lankan
    "סרי לנקה":   "סרי לנקה (צילון)",
    "צילון":      "סרי לנקה (צילון)",
    # Chinese
    "סינים":      "סין",
    # Uzbek
    "אוזבקים":    "אוזבקיסטן",
}


def _norm_country(c) -> str:
    """Normalise country value to its canonical form.

    Handles: NaN/None/empty/'-' → '' (the default key, used as fallback);
    plus a static alias map for common spelling variants (e.g. ``מודובנים``
    → ``מולדובה``). Trailing/leading whitespace is stripped.
    """
    if c is None:
        return ""
    try:
        if pd.isna(c):
            return ""
    except (ValueError, TypeError):
        pass
    s = str(c).strip()
    if s in ("", "-"):
        return ""
    return _COUNTRY_ALIASES.get(s, s)


def site_billing_lookup(standards: pd.DataFrame) -> dict:
    """
    מחזיר ``{(client_full, site, country): rule_dict}``.

    Country-aware: ``תקן.xlsx`` can contain multiple rows for the same
    (client, site) that differ ONLY by ``country`` — e.g. ``ולפמן/ולפמן``
    is ₪70 default but ₪85 for ``מודובנים`` (Moldovan workers). Previously
    we keyed the lookup by ``(client, site)`` alone, so the second row
    silently overwrote the first and **every** worker at the site got the
    Moldovan rate.

    The country key uses normalised values; the empty string ``""`` means
    "no country specified" (the default rule). ``כל האתרים`` is preserved
    as a per-client fallback rule.
    """
    lookup: dict = {}
    for _, row in standards[standards["site"] == "כל האתרים"].iterrows():
        lookup[(row["client_full"], "כל האתרים", _norm_country(row.get("country")))] = row.to_dict()
    for _, row in standards[standards["site"] != "כל האתרים"].iterrows():
        lookup[(row["client_full"], row["site"], _norm_country(row.get("country")))] = row.to_dict()
    return lookup


def get_billing_rule(
    client_full: str,
    site: str,
    lookup: dict,
    country: str | None = None,
) -> dict | None:
    """Resolve a billing rule with country priority.

    Lookup order (first match wins):
      1. ``(client, site, country)`` — exact match including country
      2. ``(client, site, "")``      — same site, no-country default
      3. ``(client, "כל האתרים", country)`` — country-specific catch-all
      4. ``(client, "כל האתרים", "")``      — generic catch-all

    The ``country`` parameter is optional for back-compat (callers that
    don't pass it get the no-country default).
    """
    c = _norm_country(country)
    # 1. site + country
    if c and (client_full, site, c) in lookup:
        return lookup[(client_full, site, c)]
    # 2. site (no country)
    if (client_full, site, "") in lookup:
        return lookup[(client_full, site, "")]
    # 3. wildcard site + country
    if c and (client_full, "כל האתרים", c) in lookup:
        return lookup[(client_full, "כל האתרים", c)]
    # 4. wildcard site (no country)
    if (client_full, "כל האתרים", "") in lookup:
        return lookup[(client_full, "כל האתרים", "")]
    return None


def get_client_billing_kind(standards: pd.DataFrame) -> dict[str, str]:
    """
    מחזיר {client_full: primary_billing_kind}.
    עדיפות: daily > hourly_with_completion > hourly_no_completion > unknown.
    """
    if standards.empty or "client_full" not in standards.columns:
        return {}

    _PRIORITY = {
        "mixed": 6, "daily_with_ot": 5, "daily_or_monthly_min": 4,
        "daily_min_only": 3, "daily_no_ot": 3,
        "hourly_with_completion": 2, "hourly_no_completion": 1,
        "unknown": 0, "missing_data": 0,
    }

    result: dict[str, str] = {}
    for client, grp in standards.groupby("client_full"):
        complete = grp[grp["is_complete"]]
        kinds    = complete["billing_kind"].tolist() if not complete.empty else grp["billing_kind"].tolist()
        best     = max(kinds, key=lambda k: _PRIORITY.get(k, 0))
        result[str(client)] = best
    return result


def enrich_from_cache(
    processed_parquet: Path | str,
    audit_parquet: Path | str,
) -> pd.DataFrame:
    """Fallback: בונה טבלת standards מ-parquet הקיים כשהקובץ לא נגיש."""
    audit_parquet = Path(audit_parquet)
    if not audit_parquet.exists():
        return pd.DataFrame()

    aud = pd.read_parquet(audit_parquet)
    rows = []
    for _, r in aud.iterrows():
        bt   = str(r.get("billing_type", "unknown") or "unknown")
        kind = _BILLING_TYPE_MAP.get(bt, "unknown")
        if kind == "daily_with_ot":
            sh = r.get("std_hours_month")
            if pd.notna(sh) and float(sh) > 0:
                kind = "daily_or_monthly_min"
        rows.append({
            "account_code":            "",
            "client_full":             str(r.get("client", "")),
            "site":                    str(r.get("site", "")),
            "country":                 "",
            "billing_type":            bt,
            "rate":                    float(r.get("hourly_rate", 0) or 0) or None,
            "include_breaks":          bool(r.get("include_breaks", False)),
            "completion_raw":          "",
            "billing_kind":            kind,
            "completion_target_hours": int(r["std_hours_month"]) if pd.notna(r.get("std_hours_month")) else None,
            "daily_min_hours":         10 if "daily" in kind else None,
            "ot_hourly_rate":          float(r.get("hourly_rate", 0) or 0) or None,
            "is_complete":             True,
        })
    return pd.DataFrame(rows)
