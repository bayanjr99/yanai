"""
Debug Writer — Phase 1.

Generates debug_full_YYYYMMDD_HHMMSS.xlsx with complete daily-level detail
for every worked row, including agreement resolution and billing breakdown.

Columns
-------
employee_id, employee_name, date, site, country,
hours_to_pay, break_hours, include_breaks,
billable_hours, ot_hours, completion_day,
rate, ot_rate, billing_type, agreement_used,
billing_amount, blocked, block_reason
"""

from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_COLS = [
    ("employee_id",   "מס' עובד"),
    ("employee_name", "שם עובד"),
    ("date",          "תאריך"),
    ("site",          "אתר"),
    ("country",       "מדינה"),
    ("hours_to_pay",  "שעות לתשלום"),
    ("break_hours",   "שעות הפסקה"),
    ("include_breaks","כולל הפסקות"),
    ("billable_hours","שעות לחישוב"),
    ("ot_hours",      "שעות נוספות"),
    ("completion_day","השלמה יומית"),
    ("rate",          "תעריף"),
    ("ot_rate",       "תעריף שע\"נ"),
    ("billing_type",  "סוג חיוב"),
    ("agreement_used","הסכם שנמצא"),
    ("billing_amount","חיוב יומי ₪"),
    ("blocked",       "חסום"),
    ("block_reason",  "סיבת חסימה"),
]

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF3FB")
_BLK_FILL = PatternFill("solid", fgColor="FFD7D7")


def write_debug(
    daily_detail_rows: list[dict],
    validation_results: list[dict],
    output_dir: str,
) -> str:
    """
    Write debug_full_*.xlsx and return its path.

    Parameters
    ----------
    daily_detail_rows : list of dicts from _bill_daily()
    validation_results: list of dicts from validation.results_to_dicts()
    output_dir        : folder to save the file
    """
    os.makedirs(output_dir, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"debug_full_{ts}.xlsx")

    wb = Workbook()

    # ── Sheet 1: daily detail ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "פירוט יומי מלא"
    _write_daily(ws1, daily_detail_rows)

    # ── Sheet 2: validation ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("אימות PDF")
    _write_validation(ws2, validation_results)

    wb.save(path)
    return path


# ─── sheet writers ────────────────────────────────────────────────────────────

def _write_daily(ws, rows: list[dict]) -> None:
    ws.cell(1, 1).value = "פירוט יומי מלא — כולל חישוב חיוב ופרטי הסכם"
    ws.cell(1, 1).font  = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLS))
    ws.append([])  # spacer

    keys    = [k for k, _ in _COLS]
    headers = [h for _, h in _COLS]

    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(3, ci)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 4):
        blocked = row.get("blocked", False)
        for ci, key in enumerate(keys, 1):
            val  = row.get(key, "")
            if isinstance(val, bool):
                val = "כן" if val else ""
            elif hasattr(val, "isoformat"):
                val = val.isoformat()
            cell = ws.cell(ri, ci)
            cell.value     = val
            cell.alignment = Alignment(horizontal="right")
            if blocked:
                cell.fill = _BLK_FILL
            elif ri % 2 == 0:
                cell.fill = _ALT_FILL

            if key in ("billing_amount", "rate", "ot_rate"):
                cell.number_format = '#,##0.00'
            elif key in ("hours_to_pay", "break_hours", "billable_hours",
                         "ot_hours", "completion_day"):
                cell.number_format = '0.00'

    # auto width
    for ci, (_, hdr) in enumerate(_COLS, 1):
        vals = [str(r.get(_COLS[ci-1][0], "")) for r in rows[:200]]
        w = max(len(hdr), max((len(v) for v in vals), default=0)) + 3
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 40)

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A4"


def _write_validation(ws, rows: list[dict]) -> None:
    ws.cell(1, 1).value = "אימות שעות עובד מול PDF"
    ws.cell(1, 1).font  = Font(bold=True, size=12, color="1F497D")
    ws.append([])

    if not rows:
        ws.cell(3, 1).value = "(אין נתונים)"
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(3, ci)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")

    _STATUS_FILLS = {
        "PASS":     PatternFill("solid", fgColor="D5F5D5"),
        "EXPECTED": PatternFill("solid", fgColor="FFF2CC"),
        "WARN":     PatternFill("solid", fgColor="FFF2CC"),
        "FAIL":     PatternFill("solid", fgColor="FFD7D7"),
    }

    for ri, row in enumerate(rows, 4):
        status = str(row.get("סטטוס", ""))
        fill   = _STATUS_FILLS.get(status)
        for ci, key in enumerate(headers, 1):
            cell = ws.cell(ri, ci)
            cell.value     = row.get(key, "")
            cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(len(h) + 6, 35)

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A4"
