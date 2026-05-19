"""
core/income_loader.py — קורא קבצי income.xlsx מתיקיות החודשים.

9 סוגי שורות:
  daily_hours      — חיוב יומי (כמות=ימים, מחיר/יחידה > 200₪)
  hourly_hours     — חיוב שעתי (כמות=שעות, מחיר/יחידה <= 200₪)
  completion_hours — השלמה לתקן (פריט 1111 + "השלמה")
  overtime_hours   — שעות נוספות מעל 10/יום (פריט 1111 + "שעות נוספות")
  supplement_hours — תוסף 1111 אחר
  housing          — מיגורים (פריט 2 / "מיגור")
  settlement       — התחשבנות (פריט 16 / "התחשבנות")
  import_fee       — השתתפות יבוא עובד (פריט 17 / "השתתפות")
  fee_refund       — החזר אגרות ("אגרות" / "אגרת")
  credit           — זיכוי ("זיכוי") — מוצג כסכום שלילי
  other            — כל השאר
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

_MONTH_RE = re.compile(r"^\d{2}-\d{4}$")

_ALL_KINDS = [
    "daily_hours", "hourly_hours", "completion_hours", "overtime_hours",
    "supplement_hours", "housing", "settlement", "import_fee",
    "fee_refund", "credit", "anomaly", "other",
]


def _classify_item(item_num, description: str, qty: float, amount: float) -> str:
    """מסווג שורת פרט לאחד מ-12 סוגים."""
    s    = str(item_num).strip()
    desc = str(description)

    # זיכוי אמיתי = סכום שלילי, או תיאור "זיכוי" עם כמות/סכום שליליים
    # חשוב: "זיכוי" בתיאור עם qty>0 ו-amount>0 = תוספת חיוב (כגון "זיכוי עבור הפרש מחיר")
    if qty < 0 and amount < 0:
        return "credit"
    if "זיכוי" in desc and (qty <= 0 or amount <= 0):
        return "credit"

    # שורה חריגה: כמות וסכום בסימנים מנוגדים
    if (qty < 0 and amount > 0) or (qty > 0 and amount < 0):
        return "anomaly"

    # item=1 ו-1111 — מספר הפריט קודם לבדיקת תיאור (גם אם יש "אגרות" בתיאור)
    if s == "1":
        if qty > 0 and amount > 0:
            unit_price = amount / qty
            return "daily_hours" if unit_price > 200 else "hourly_hours"
        return "hourly_hours"

    if s == "1111":
        if "השלמה" in desc:
            return "completion_hours"
        if "שעות נוספות" in desc or "מעל 10" in desc:
            return "overtime_hours"
        return "supplement_hours"

    # בדיקות לפי תיאור — רק לפריטים שאינם 1/1111
    if s == "2" or "מיגור" in desc:
        return "housing"

    if s == "16" or "התחשבנות" in desc:
        return "settlement"

    if s == "17" or "השתתפות" in desc:
        return "import_fee"

    if "אגרות" in desc or "אגרת" in desc:
        return "fee_refund"

    return "other"


def load_income_files(data_dir: Path | str = "data") -> pd.DataFrame:
    """
    קורא את כל קבצי income.xlsx ומחזיר DataFrame מאוגר לפי (month, client_full).

    עמודות פלט עיקריות:
      billing_amount   — סך הכנסה (כולל זיכויים שליליים)
      billed_hours     — שעות שחויבו (שעתי + השלמה + נוספות + supplement)
      billed_days      — ימים שחויבו (יומי)
      billing_work     — הכנסה מעבודה (שעתי+יומי+השלמות)
      billing_extras   — תוספות (מיגור+התחשבנות+יבוא+אגרות)
      billing_credits  — זיכויים (ערך שלילי)
      billing_type_actual — "hourly" | "daily" | "mixed" | "none"
      amount_*/qty_*   — פירוט לכל סוג
    """
    data_dir = Path(data_dir)
    rows: list[dict] = []

    for month_dir in sorted(data_dir.iterdir()):
        if not month_dir.is_dir():
            continue
        if not _MONTH_RE.match(month_dir.name):
            continue
        income_file = month_dir / "income.xlsx"
        if not income_file.exists():
            continue

        try:
            wb = openpyxl.load_workbook(str(income_file), data_only=True)
        except Exception as exc:
            print(f"[income] WARNING: Cannot open {income_file}: {exc}")
            continue

        ws = wb[wb.sheetnames[0]]
        current_client: str | None = None

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 6:
                continue
            a, b, c, d, e, f = (row[i] if i < len(row) else None for i in range(6))

            if a == "שם חשבון" and b:
                current_client = str(b).strip()
                continue

            if current_client is None:
                continue

            if a and "סה" in str(a) and "מפתח" in str(a):
                continue

            if c is None or d is None:
                continue
            try:
                int(str(c).strip())
            except (ValueError, TypeError):
                continue

            qty  = float(e) if isinstance(e, (int, float)) else 0.0
            amt  = float(f) if isinstance(f, (int, float)) else 0.0
            desc = str(d).strip()
            kind = _classify_item(c, desc, qty, amt)

            # שמור ערכים כפי שהם בקובץ — ללא היפוך מלאכותי
            rows.append({
                "month":       month_dir.name,
                "client_full": current_client,
                "description": desc,
                "qty":         qty,
                "amount":      amt,
                "kind":        kind,
            })

    if not rows:
        return pd.DataFrame()

    detail = pd.DataFrame(rows)
    agg_rows: list[dict] = []

    for (month, client), g in detail.groupby(["month", "client_full"]):
        rec: dict = {"month": month, "client_full": client}

        for kind in _ALL_KINDS:
            sub = g[g["kind"] == kind]
            rec[f"qty_{kind}"]    = float(sub["qty"].sum())
            rec[f"amount_{kind}"] = float(sub["amount"].sum())

        _HOUR_KINDS = {"hourly_hours", "completion_hours", "overtime_hours", "supplement_hours"}
        rec["billed_hours"] = sum(rec[f"qty_{k}"] for k in _HOUR_KINDS)
        rec["billed_days"]  = rec["qty_daily_hours"]

        rec["billing_amount"] = float(g["amount"].sum())

        rec["billing_work"] = (
            rec["amount_hourly_hours"] + rec["amount_daily_hours"] +
            rec["amount_completion_hours"] + rec["amount_overtime_hours"] +
            rec["amount_supplement_hours"]
        )
        rec["billing_extras"] = (
            rec["amount_housing"] + rec["amount_settlement"] +
            rec["amount_import_fee"] + rec["amount_fee_refund"]
        )
        rec["billing_credits"] = rec["amount_credit"]  # שלילי

        has_daily  = rec["qty_daily_hours"]  > 0
        has_hourly = rec["qty_hourly_hours"] > 0
        if has_daily and not has_hourly:
            rec["billing_type_actual"] = "daily"
        elif has_hourly and not has_daily:
            rec["billing_type_actual"] = "hourly"
        elif has_daily and has_hourly:
            rec["billing_type_actual"] = "mixed"
        else:
            rec["billing_type_actual"] = "none"

        agg_rows.append(rec)

    return pd.DataFrame(agg_rows)


def aggregate_to_month_client(income_df: pd.DataFrame) -> pd.DataFrame:
    """כבר אגור — מחזיר אותו."""
    return income_df
